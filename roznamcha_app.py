import hashlib
import hmac
import html
import os
import secrets
import sqlite3
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from io import BytesIO
from textwrap import dedent

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill


# =========================================================
# APPLICATION CONFIGURATION
# =========================================================

APP_TITLE = "Multi-School Roznamcha Attendance System"
APP_VERSION = "Multi-School Version 1.0 — 01 September 2026"

DB_PATH = os.getenv(
    "ROZNAMCHA_DB",
    "multi_school_roznamcha.db",
)

DEFAULT_CLASSES = [
    "Katchi",
    "1",
    "2",
    "3",
    "4",
    "5",
]

# Change this in Streamlit Cloud Secrets before publishing.
DEFAULT_MASTER_REGISTRATION_PASSWORD = "AddSchool@2026!"

st.set_page_config(
    page_title=APP_TITLE,
    page_icon="📘",
    layout="wide",
)


def get_setting(name, default):
    try:
        if name in st.secrets:
            return str(st.secrets[name])
    except (FileNotFoundError, KeyError):
        pass

    return os.getenv(name, default)


MASTER_REGISTRATION_PASSWORD = get_setting(
    "MASTER_REGISTRATION_PASSWORD",
    DEFAULT_MASTER_REGISTRATION_PASSWORD,
)


# =========================================================
# RESPONSIVE CSS
# =========================================================

st.markdown(
    dedent(
        """
        <style>
        .block-container {
            padding-top: 1.35rem;
            padding-bottom: 2rem;
            max-width: 1600px;
        }

        div[data-testid="stMarkdownContainer"],
        div[data-testid="stMarkdownContainer"] > div {
            height: auto !important;
            max-height: none !important;
            overflow: visible !important;
        }

        .stApp,
        .stApp button,
        .stApp input,
        .stApp textarea,
        .stApp label,
        .stApp p {
            font-family: Arial, Helvetica, sans-serif !important;
        }

        .page-heading {
            display: flex;
            align-items: center;
            gap: 14px;
            width: 100%;
            height: auto !important;
            min-height: 76px;
            overflow: visible !important;
            margin: 6px 0 20px 0;
            padding: 6px 0;
        }

        .page-title {
            display: block;
            width: 100%;
            height: auto !important;
            min-height: 1.55em;
            overflow: visible !important;
            white-space: normal !important;
            overflow-wrap: anywhere;
            color: #242938;
            font-family: Arial, Helvetica, sans-serif !important;
            font-size: clamp(30px, 4vw, 43px) !important;
            font-weight: 700 !important;
            line-height: 1.5 !important;
            margin: 0 !important;
            padding: 4px 0 8px 0 !important;
        }

        .page-subtitle {
            display: block;
            height: auto !important;
            overflow: visible !important;
            white-space: normal !important;
            color: #657080;
            font-family: Arial, Helvetica, sans-serif !important;
            font-size: 16px;
            line-height: 1.6 !important;
            margin: 2px 0 0 0 !important;
        }

        .brand-badge {
            display: flex;
            align-items: center;
            justify-content: center;
            flex: 0 0 52px;
            width: 52px;
            height: 52px;
            border: 3px solid #08243c;
            border-radius: 8px;
            background: #1688c9;
            color: #ffffff;
            font-family: Arial, Helvetica, sans-serif !important;
            font-size: 20px;
            font-weight: 700;
            line-height: 1;
        }

        .section-title {
            display: block;
            width: 100%;
            height: auto !important;
            min-height: 1.65em;
            overflow: visible !important;
            white-space: normal !important;
            color: #2b3040;
            font-family: Arial, Helvetica, sans-serif !important;
            font-size: clamp(22px, 2.5vw, 28px);
            font-weight: 700 !important;
            line-height: 1.5 !important;
            margin: 20px 0 10px 0;
            padding: 4px 0 8px 0;
        }

        .form-title {
            display: block;
            width: 100%;
            height: auto !important;
            min-height: 1.65em;
            overflow: visible !important;
            white-space: normal !important;
            color: #303646;
            font-family: Arial, Helvetica, sans-serif !important;
            font-size: clamp(18px, 2vw, 22px);
            font-weight: 700 !important;
            line-height: 1.5 !important;
            margin: 2px 0 10px 0;
            padding: 4px 0 8px 0;
        }

        .sidebar-brand {
            color: #242938;
            font-family: Arial, Helvetica, sans-serif !important;
            font-size: 24px;
            font-weight: 700;
            line-height: 1.5;
            min-height: 1.5em;
            overflow: visible !important;
            padding: 4px 0;
        }

        .dashboard-heading {
            display: block;
            width: 100%;
            height: auto !important;
            overflow: visible !important;
            background: linear-gradient(90deg, #071426, #102b49);
            border: 1px solid #274766;
            border-radius: 12px;
            padding: 18px 22px;
            margin: 8px 0 18px 0;
        }

        .dashboard-title {
            display: block;
            height: auto !important;
            min-height: 1.65em;
            overflow: visible !important;
            white-space: normal !important;
            color: #ffffff;
            font-family: Arial, Helvetica, sans-serif !important;
            font-size: clamp(23px, 3vw, 32px) !important;
            font-weight: 700 !important;
            line-height: 1.5 !important;
            margin: 0 !important;
            padding: 3px 0 6px 0 !important;
        }

        .dashboard-subtitle {
            color: #b9d8f4;
            font-family: Arial, Helvetica, sans-serif !important;
            font-size: 16px;
            line-height: 1.55 !important;
        }

        .rozn-scroll {
            width: 100%;
            overflow-x: auto;
            border-radius: 8px;
            margin: 8px 0 16px 0;
        }

        table.rozn-table {
            width: 100%;
            min-width: 980px;
            border-collapse: collapse;
            background: #030303;
            color: #ffffff;
            text-align: center;
            font-family: Arial, Helvetica, sans-serif;
        }

        .rozn-table th,
        .rozn-table td {
            border: 1px solid #aeb4bb;
            padding: 8px 9px;
            font-size: 16px;
            white-space: nowrap;
        }

        .rozn-table thead th {
            font-size: 17px;
            font-weight: 800;
        }

        .class-head,
        .strength-head,
        .strength-cell,
        .class-cell {
            background: #050505;
        }

        .class-cell {
            text-align: left;
            color: #ffffff;
        }

        .present-head {
            background: #00df20;
            color: #061b08;
        }

        .present-cell {
            background: #00c921;
            color: #ffffff;
        }

        .absent-head {
            background: #ec1010;
            color: #ffffff;
        }

        .absent-cell {
            background: #671b1b;
            color: #ffffff;
        }

        .percentage-head {
            background: #050505;
            color: #2fe9ff;
        }

        .percentage-cell {
            background: #1e5fae;
            color: #ffffff;
        }

        .total-row td {
            font-size: 17px;
            font-weight: 800;
        }

        @media (max-width: 700px) {
            .page-title {
                font-size: 28px !important;
            }

            .brand-badge {
                flex-basis: 44px;
                width: 44px;
                height: 44px;
                font-size: 17px;
            }

            .dashboard-heading {
                padding: 14px 16px;
            }

            .dashboard-title {
                font-size: 24px !important;
            }
        }
        </style>
        """
    ),
    unsafe_allow_html=True,
)


# =========================================================
# SAFE HEADINGS
# =========================================================

def page_heading(
    title,
    subtitle="",
    show_badge=False,
):
    badge = (
        '<div class="brand-badge">MS</div>'
        if show_badge
        else ""
    )

    subtitle_html = (
        f'<div class="page-subtitle">'
        f'{html.escape(subtitle)}'
        f'</div>'
        if subtitle
        else ""
    )

    content = (
        '<div class="page-heading">'
        f'{badge}'
        '<div style="display:block;width:100%;'
        'height:auto;overflow:visible;">'
        f'<div class="page-title">'
        f'{html.escape(title)}'
        f'</div>'
        f'{subtitle_html}'
        '</div>'
        '</div>'
    )

    st.markdown(
        content,
        unsafe_allow_html=True,
    )


def section_heading(title):
    st.markdown(
        f'<div class="section-title">'
        f'{html.escape(title)}'
        f'</div>',
        unsafe_allow_html=True,
    )


def form_heading(title):
    st.markdown(
        f'<div class="form-title">'
        f'{html.escape(title)}'
        f'</div>',
        unsafe_allow_html=True,
    )


# =========================================================
# DATABASE AND PASSWORD SECURITY
# =========================================================

@contextmanager
def db_connection():
    connection = sqlite3.connect(
        DB_PATH,
        timeout=30,
    )

    connection.row_factory = sqlite3.Row

    connection.execute(
        "PRAGMA busy_timeout = 30000"
    )

    connection.execute(
        "PRAGMA foreign_keys = ON"
    )

    connection.execute(
        "PRAGMA journal_mode = WAL"
    )

    try:
        yield connection
        connection.commit()

    except Exception:
        connection.rollback()
        raise

    finally:
        connection.close()


def hash_password(password):
    salt = secrets.token_bytes(16)

    digest = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt,
        200_000,
    )

    return (
        f"{salt.hex()}$"
        f"{digest.hex()}"
    )


def verify_password(
    password,
    stored_password,
):
    try:
        salt_hex, stored_digest = (
            stored_password.split("$", 1)
        )

        calculated = hashlib.pbkdf2_hmac(
            "sha256",
            password.encode("utf-8"),
            bytes.fromhex(salt_hex),
            200_000,
        ).hex()

        return hmac.compare_digest(
            calculated,
            stored_digest,
        )

    except (ValueError, AttributeError):
        return False


def initialize_database():
    with db_connection() as con:
        con.executescript(
            """
            CREATE TABLE IF NOT EXISTS ms_schools (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                school_name TEXT NOT NULL,
                username TEXT NOT NULL
                    UNIQUE COLLATE NOCASE,
                password_hash TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS ms_classes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                school_id INTEGER NOT NULL,
                class_name TEXT NOT NULL COLLATE NOCASE,
                sort_order INTEGER NOT NULL DEFAULT 0,
                boys_enabled INTEGER NOT NULL DEFAULT 1,
                girls_enabled INTEGER NOT NULL DEFAULT 1,
                boys_strength INTEGER NOT NULL DEFAULT 0
                    CHECK(boys_strength >= 0),
                girls_strength INTEGER NOT NULL DEFAULT 0
                    CHECK(girls_strength >= 0),
                active INTEGER NOT NULL DEFAULT 1,

                UNIQUE(school_id, class_name),

                FOREIGN KEY(school_id)
                    REFERENCES ms_schools(id)
                    ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS ms_attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                school_id INTEGER NOT NULL,
                class_id INTEGER NOT NULL,
                attendance_date TEXT NOT NULL,
                boys_strength INTEGER NOT NULL DEFAULT 0,
                girls_strength INTEGER NOT NULL DEFAULT 0,
                boys_present INTEGER NOT NULL DEFAULT 0,
                girls_present INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,

                UNIQUE(
                    school_id,
                    class_id,
                    attendance_date
                ),

                FOREIGN KEY(school_id)
                    REFERENCES ms_schools(id)
                    ON DELETE CASCADE,

                FOREIGN KEY(class_id)
                    REFERENCES ms_classes(id)
                    ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS
            idx_ms_attendance_school_date
            ON ms_attendance(
                school_id,
                attendance_date
            );

            CREATE INDEX IF NOT EXISTS
            idx_ms_classes_school
            ON ms_classes(
                school_id,
                active,
                sort_order
            );
            """
        )


# =========================================================
# SCHOOL ACCOUNT FUNCTIONS
# =========================================================

def create_school_account(
    master_password,
    school_name,
    username,
    password,
    confirm_password,
):
    school_name = school_name.strip()
    username = username.strip()

    if not hmac.compare_digest(
        master_password,
        MASTER_REGISTRATION_PASSWORD,
    ):
        return (
            False,
            "The master registration password is incorrect.",
        )

    if len(school_name) < 3:
        return (
            False,
            "Enter a valid school name.",
        )

    if len(username) < 3:
        return (
            False,
            "Username must contain at least 3 characters.",
        )

    if " " in username:
        return (
            False,
            "Username cannot contain spaces.",
        )

    if len(password) < 8:
        return (
            False,
            "Password must contain at least 8 characters.",
        )

    if password != confirm_password:
        return (
            False,
            "The passwords do not match.",
        )

    timestamp = (
        datetime.now()
        .replace(microsecond=0)
        .isoformat()
    )

    try:
        with db_connection() as con:
            cursor = con.execute(
                """
                INSERT INTO ms_schools(
                    school_name,
                    username,
                    password_hash,
                    active,
                    created_at
                )
                VALUES (?, ?, ?, 1, ?)
                """,
                (
                    school_name,
                    username,
                    hash_password(password),
                    timestamp,
                ),
            )

            school_id = cursor.lastrowid

            for order, class_name in enumerate(
                DEFAULT_CLASSES,
                start=1,
            ):
                con.execute(
                    """
                    INSERT INTO ms_classes(
                        school_id,
                        class_name,
                        sort_order,
                        boys_enabled,
                        girls_enabled,
                        boys_strength,
                        girls_strength,
                        active
                    )
                    VALUES (?, ?, ?, 1, 1, 0, 0, 1)
                    """,
                    (
                        school_id,
                        class_name,
                        order,
                    ),
                )

        return (
            True,
            "School account created. You can now sign in.",
        )

    except sqlite3.IntegrityError:
        return (
            False,
            "This username is already being used "
            "by another school.",
        )


def authenticate(
    username,
    password,
):
    with db_connection() as con:
        school = con.execute(
            """
            SELECT
                id,
                school_name,
                username,
                password_hash
            FROM ms_schools
            WHERE username = ?
              AND active = 1
            """,
            (username.strip(),),
        ).fetchone()

    if (
        school
        and verify_password(
            password,
            school["password_hash"],
        )
    ):
        return {
            "school_id": school["id"],
            "school_name": school["school_name"],
            "username": school["username"],
        }

    return None


def get_school(school_id):
    with db_connection() as con:
        row = con.execute(
            """
            SELECT
                id,
                school_name,
                username
            FROM ms_schools
            WHERE id = ?
              AND active = 1
            """,
            (school_id,),
        ).fetchone()

    return dict(row) if row else None


def update_school_name(
    school_id,
    new_name,
):
    new_name = new_name.strip()

    if len(new_name) < 3:
        return (
            False,
            "Enter a valid school name.",
        )

    with db_connection() as con:
        cursor = con.execute(
            """
            UPDATE ms_schools
            SET school_name = ?
            WHERE id = ?
              AND active = 1
            """,
            (
                new_name,
                school_id,
            ),
        )

        if cursor.rowcount != 1:
            return (
                False,
                "School account was not found.",
            )

    return (
        True,
        "School name updated successfully.",
    )


def change_username(
    school_id,
    current_password,
    new_username,
):
    new_username = new_username.strip()

    if len(new_username) < 3:
        return (
            False,
            "Username must contain at least 3 characters.",
        )

    if " " in new_username:
        return (
            False,
            "Username cannot contain spaces.",
        )

    with db_connection() as con:
        row = con.execute(
            """
            SELECT password_hash
            FROM ms_schools
            WHERE id = ?
              AND active = 1
            """,
            (school_id,),
        ).fetchone()

        if (
            not row
            or not verify_password(
                current_password,
                row["password_hash"],
            )
        ):
            return (
                False,
                "Current password is incorrect.",
            )

        try:
            con.execute(
                """
                UPDATE ms_schools
                SET username = ?
                WHERE id = ?
                """,
                (
                    new_username,
                    school_id,
                ),
            )

        except sqlite3.IntegrityError:
            return (
                False,
                "This username is already being used "
                "by another school.",
            )

    return (
        True,
        "Username changed successfully.",
    )


def change_password(
    school_id,
    current_password,
    new_password,
):
    if len(new_password) < 8:
        return (
            False,
            "New password must contain at least 8 characters.",
        )

    with db_connection() as con:
        row = con.execute(
            """
            SELECT password_hash
            FROM ms_schools
            WHERE id = ?
              AND active = 1
            """,
            (school_id,),
        ).fetchone()

        if (
            not row
            or not verify_password(
                current_password,
                row["password_hash"],
            )
        ):
            return (
                False,
                "Current password is incorrect.",
            )

        con.execute(
            """
            UPDATE ms_schools
            SET password_hash = ?
            WHERE id = ?
            """,
            (
                hash_password(new_password),
                school_id,
            ),
        )

    return (
        True,
        "Password changed successfully.",
    )


# =========================================================
# SCHOOL-SPECIFIC CLASS FUNCTIONS
# =========================================================

def get_classes(
    school_id,
    active_only=True,
):
    active_condition = (
        "AND active = 1"
        if active_only
        else ""
    )

    with db_connection() as con:
        classes = con.execute(
            f"""
            SELECT *
            FROM ms_classes
            WHERE school_id = ?
            {active_condition}
            ORDER BY
                sort_order,
                class_name
            """,
            (school_id,),
        ).fetchall()

    return classes


def update_class(
    school_id,
    class_id,
    class_name,
    sort_order,
    boys_enabled,
    girls_enabled,
    boys_strength,
    girls_strength,
    active,
):
    class_name = class_name.strip()

    if not class_name:
        return (
            False,
            "Class name cannot be empty.",
        )

    if not boys_enabled and not girls_enabled:
        return (
            False,
            "Enable at least one section.",
        )

    try:
        with db_connection() as con:
            cursor = con.execute(
                """
                UPDATE ms_classes
                SET
                    class_name = ?,
                    sort_order = ?,
                    boys_enabled = ?,
                    girls_enabled = ?,
                    boys_strength = ?,
                    girls_strength = ?,
                    active = ?
                WHERE id = ?
                  AND school_id = ?
                """,
                (
                    class_name,
                    int(sort_order),
                    int(boys_enabled),
                    int(girls_enabled),
                    (
                        int(boys_strength)
                        if boys_enabled
                        else 0
                    ),
                    (
                        int(girls_strength)
                        if girls_enabled
                        else 0
                    ),
                    int(active),
                    class_id,
                    school_id,
                ),
            )

            if cursor.rowcount != 1:
                return (
                    False,
                    "This class does not belong "
                    "to your school.",
                )

        return (
            True,
            "Class configuration updated successfully.",
        )

    except sqlite3.IntegrityError:
        return (
            False,
            "Your school already has a class "
            "with this name.",
        )


def add_class(
    school_id,
    class_name,
    sort_order,
    boys_enabled,
    girls_enabled,
    boys_strength,
    girls_strength,
):
    class_name = class_name.strip()

    if not class_name:
        return (
            False,
            "Enter a class name.",
        )

    if not boys_enabled and not girls_enabled:
        return (
            False,
            "Enable at least one section.",
        )

    try:
        with db_connection() as con:
            con.execute(
                """
                INSERT INTO ms_classes(
                    school_id,
                    class_name,
                    sort_order,
                    boys_enabled,
                    girls_enabled,
                    boys_strength,
                    girls_strength,
                    active
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, 1)
                """,
                (
                    school_id,
                    class_name,
                    int(sort_order),
                    int(boys_enabled),
                    int(girls_enabled),
                    (
                        int(boys_strength)
                        if boys_enabled
                        else 0
                    ),
                    (
                        int(girls_strength)
                        if girls_enabled
                        else 0
                    ),
                ),
            )

        return (
            True,
            "New class added successfully.",
        )

    except sqlite3.IntegrityError:
        return (
            False,
            "Your school already has a class "
            "with this name.",
        )


# =========================================================
# SCHOOL-SPECIFIC ATTENDANCE FUNCTIONS
# =========================================================

def get_existing_attendance(
    school_id,
    attendance_date,
):
    with db_connection() as con:
        rows = con.execute(
            """
            SELECT
                class_id,
                boys_present,
                girls_present
            FROM ms_attendance
            WHERE school_id = ?
              AND attendance_date = ?
            """,
            (
                school_id,
                attendance_date.isoformat(),
            ),
        ).fetchall()

    return {
        row["class_id"]: dict(row)
        for row in rows
    }


def save_attendance(
    school_id,
    selected_date,
    entries,
):
    timestamp = (
        datetime.now()
        .replace(microsecond=0)
        .isoformat()
    )

    with db_connection() as con:
        allowed_class_ids = {
            row["id"]
            for row in con.execute(
                """
                SELECT id
                FROM ms_classes
                WHERE school_id = ?
                  AND active = 1
                """,
                (school_id,),
            ).fetchall()
        }

        for entry in entries:
            if (
                entry["class_id"]
                not in allowed_class_ids
            ):
                raise ValueError(
                    "A class does not belong to "
                    "the logged-in school."
                )

            con.execute(
                """
                INSERT INTO ms_attendance(
                    school_id,
                    class_id,
                    attendance_date,
                    boys_strength,
                    girls_strength,
                    boys_present,
                    girls_present,
                    created_at,
                    updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)

                ON CONFLICT(
                    school_id,
                    class_id,
                    attendance_date
                )
                DO UPDATE SET
                    boys_strength =
                        excluded.boys_strength,
                    girls_strength =
                        excluded.girls_strength,
                    boys_present =
                        excluded.boys_present,
                    girls_present =
                        excluded.girls_present,
                    updated_at =
                        excluded.updated_at
                """,
                (
                    school_id,
                    entry["class_id"],
                    selected_date.isoformat(),
                    entry["boys_strength"],
                    entry["girls_strength"],
                    entry["boys_present"],
                    entry["girls_present"],
                    timestamp,
                    timestamp,
                ),
            )


def attendance_dataframe(
    school_id,
    start_date,
    end_date,
):
    with db_connection() as con:
        data = pd.read_sql_query(
            """
            SELECT
                a.attendance_date AS Date,
                c.class_name AS Class,
                c.sort_order AS Class_Order,
                a.boys_strength AS Boys_Strength,
                a.girls_strength AS Girls_Strength,
                a.boys_present AS Boys_Present,
                a.girls_present AS Girls_Present
            FROM ms_attendance AS a

            JOIN ms_classes AS c
              ON c.id = a.class_id
             AND c.school_id = a.school_id

            WHERE a.school_id = ?
              AND a.attendance_date
                  BETWEEN ? AND ?

            ORDER BY
                a.attendance_date,
                c.sort_order,
                c.class_name
            """,
            con,
            params=(
                school_id,
                start_date.isoformat(),
                end_date.isoformat(),
            ),
        )

    if data.empty:
        return data

    numeric_columns = [
        "Boys_Strength",
        "Girls_Strength",
        "Boys_Present",
        "Girls_Present",
    ]

    data[numeric_columns] = (
        data[numeric_columns]
        .fillna(0)
        .astype(int)
    )

    data["Total_Strength"] = (
        data["Boys_Strength"]
        + data["Girls_Strength"]
    )

    data["Boys_Absent"] = (
        data["Boys_Strength"]
        - data["Boys_Present"]
    )

    data["Girls_Absent"] = (
        data["Girls_Strength"]
        - data["Girls_Present"]
    )

    data["Total_Present"] = (
        data["Boys_Present"]
        + data["Girls_Present"]
    )

    data["Total_Absent"] = (
        data["Boys_Absent"]
        + data["Girls_Absent"]
    )

    data["Attendance_Percentage"] = (
        data["Total_Present"]
        .div(
            data["Total_Strength"]
            .replace(0, pd.NA)
        )
        .fillna(0)
    )

    data["Date"] = (
        pd.to_datetime(data["Date"])
        .dt.date
    )

    return data


def attendance_metrics(data):
    if data.empty:
        return 0, 0, 0, 0.0

    strength = int(
        data["Total_Strength"].sum()
    )

    present = int(
        data["Total_Present"].sum()
    )

    absent = int(
        data["Total_Absent"].sum()
    )

    percentage = (
        present / strength
        if strength
        else 0.0
    )

    return (
        strength,
        present,
        absent,
        percentage,
    )


def class_summary_dataframe(data):
    if data.empty:
        return pd.DataFrame()

    summary = (
        data.groupby(
            "Class",
            as_index=False,
        )
        .agg(
            Class_Order=(
                "Class_Order",
                "min",
            ),
            Boys_Strength=(
                "Boys_Strength",
                "sum",
            ),
            Girls_Strength=(
                "Girls_Strength",
                "sum",
            ),
            Boys_Present=(
                "Boys_Present",
                "sum",
            ),
            Girls_Present=(
                "Girls_Present",
                "sum",
            ),
        )
        .sort_values(
            [
                "Class_Order",
                "Class",
            ],
            kind="stable",
        )
    )

    summary["Total_Strength"] = (
        summary["Boys_Strength"]
        + summary["Girls_Strength"]
    )

    summary["Total_Present"] = (
        summary["Boys_Present"]
        + summary["Girls_Present"]
    )

    summary["Boys_Absent"] = (
        summary["Boys_Strength"]
        - summary["Boys_Present"]
    )

    summary["Girls_Absent"] = (
        summary["Girls_Strength"]
        - summary["Girls_Present"]
    )

    summary["Total_Absent"] = (
        summary["Boys_Absent"]
        + summary["Girls_Absent"]
    )

    summary["Attendance_Percentage"] = (
        summary["Total_Present"]
        .div(
            summary["Total_Strength"]
            .replace(0, pd.NA)
        )
        .fillna(0)
    )

    return summary


# =========================================================
# WORKBOOK-STYLE DASHBOARD TABLE
# =========================================================

def render_roznamcha_table(data):
    summary = class_summary_dataframe(data)

    if summary.empty:
        st.info(
            "Attendance has not been entered "
            "for this date."
        )
        return

    body_rows = []

    for _, row in summary.iterrows():
        body_rows.append(
            "<tr>"
            f'<td class="class-cell">'
            f'{html.escape(str(row["Class"]))}'
            "</td>"
            f'<td class="strength-cell">'
            f'{int(row["Boys_Strength"])}'
            "</td>"
            f'<td class="strength-cell">'
            f'{int(row["Girls_Strength"])}'
            "</td>"
            f'<td class="strength-cell">'
            f'{int(row["Total_Strength"])}'
            "</td>"
            f'<td class="present-cell">'
            f'{int(row["Boys_Present"])}'
            "</td>"
            f'<td class="present-cell">'
            f'{int(row["Girls_Present"])}'
            "</td>"
            f'<td class="present-cell">'
            f'{int(row["Total_Present"])}'
            "</td>"
            f'<td class="absent-cell">'
            f'{int(row["Boys_Absent"])}'
            "</td>"
            f'<td class="absent-cell">'
            f'{int(row["Girls_Absent"])}'
            "</td>"
            f'<td class="absent-cell">'
            f'{int(row["Total_Absent"])}'
            "</td>"
            f'<td class="percentage-cell">'
            f'{row["Attendance_Percentage"]:.2%}'
            "</td>"
            "</tr>"
        )

    total_fields = [
        "Boys_Strength",
        "Girls_Strength",
        "Total_Strength",
        "Boys_Present",
        "Girls_Present",
        "Total_Present",
        "Boys_Absent",
        "Girls_Absent",
        "Total_Absent",
    ]

    totals = {
        field: int(summary[field].sum())
        for field in total_fields
    }

    total_percentage = (
        totals["Total_Present"]
        / totals["Total_Strength"]
        if totals["Total_Strength"]
        else 0.0
    )

    table_html = (
        '<div class="rozn-scroll">'
        '<table class="rozn-table">'
        '<thead>'

        '<tr>'
        '<th class="class-head" rowspan="2">'
        'Classes'
        '</th>'
        '<th class="strength-head" colspan="3">'
        'Total Strength'
        '</th>'
        '<th class="present-head" colspan="3">'
        'Present'
        '</th>'
        '<th class="absent-head" colspan="3">'
        'Absentees'
        '</th>'
        '<th class="percentage-head" rowspan="2">'
        'Percentage<br>Attendance'
        '</th>'
        '</tr>'

        '<tr>'
        '<th class="strength-head">Boys</th>'
        '<th class="strength-head">Girls</th>'
        '<th class="strength-head">Total</th>'
        '<th class="present-head">Boys</th>'
        '<th class="present-head">Girls</th>'
        '<th class="present-head">Total</th>'
        '<th class="absent-head">Boys</th>'
        '<th class="absent-head">Girls</th>'
        '<th class="absent-head">Total</th>'
        '</tr>'

        '</thead>'
        '<tbody>'

        f'{"".join(body_rows)}'

        '<tr class="total-row">'
        '<td class="class-cell">Total</td>'

        f'<td class="strength-cell">'
        f'{totals["Boys_Strength"]}'
        f'</td>'

        f'<td class="strength-cell">'
        f'{totals["Girls_Strength"]}'
        f'</td>'

        f'<td class="strength-cell">'
        f'{totals["Total_Strength"]}'
        f'</td>'

        f'<td class="present-cell">'
        f'{totals["Boys_Present"]}'
        f'</td>'

        f'<td class="present-cell">'
        f'{totals["Girls_Present"]}'
        f'</td>'

        f'<td class="present-cell">'
        f'{totals["Total_Present"]}'
        f'</td>'

        f'<td class="absent-cell">'
        f'{totals["Boys_Absent"]}'
        f'</td>'

        f'<td class="absent-cell">'
        f'{totals["Girls_Absent"]}'
        f'</td>'

        f'<td class="absent-cell">'
        f'{totals["Total_Absent"]}'
        f'</td>'

        f'<td class="percentage-cell">'
        f'{total_percentage:.2%}'
        f'</td>'

        '</tr>'
        '</tbody>'
        '</table>'
        '</div>'
    )

    st.markdown(
        table_html,
        unsafe_allow_html=True,
    )


# =========================================================
# EXCEL REPORT FUNCTIONS
# =========================================================

def daily_summary_dataframe(data):
    if data.empty:
        return pd.DataFrame()

    summary = (
        data.groupby(
            "Date",
            as_index=False,
        )
        .agg(
            Attendance_Opportunities=(
                "Total_Strength",
                "sum",
            ),
            Total_Present=(
                "Total_Present",
                "sum",
            ),
            Total_Absent=(
                "Total_Absent",
                "sum",
            ),
        )
    )

    summary["Attendance_Percentage"] = (
        summary["Total_Present"]
        .div(
            summary[
                "Attendance_Opportunities"
            ].replace(0, pd.NA)
        )
        .fillna(0)
    )

    return summary


def style_excel_sheet(worksheet):
    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    for column_cells in worksheet.columns:
        values = [
            (
                str(cell.value)
                if cell.value is not None
                else ""
            )
            for cell in column_cells
        ]

        width = min(
            max(
                max(
                    map(len, values),
                    default=0,
                ) + 2,
                11,
            ),
            32,
        )

        worksheet.column_dimensions[
            column_cells[0].column_letter
        ].width = width

    for row in worksheet.iter_rows(
        min_row=2
    ):
        for cell in row:
            heading = str(
                worksheet.cell(
                    1,
                    cell.column,
                ).value
            )

            if "Percentage" in heading:
                cell.number_format = "0.00%"


def make_excel_report(
    data,
    school_name,
    title,
    start_date,
    end_date,
):
    output = BytesIO()

    details = data[
        [
            "Date",
            "Class",
            "Boys_Strength",
            "Girls_Strength",
            "Total_Strength",
            "Boys_Present",
            "Girls_Present",
            "Total_Present",
            "Boys_Absent",
            "Girls_Absent",
            "Total_Absent",
            "Attendance_Percentage",
        ]
    ].copy()

    daily = daily_summary_dataframe(data)

    (
        strength,
        present,
        absent,
        percentage,
    ) = attendance_metrics(data)

    overview = pd.DataFrame(
        {
            "Item": [
                "Report",
                "School",
                "Start Date",
                "End Date",
                "Attendance Opportunities",
                "Present",
                "Absent",
                "Attendance Percentage",
            ],
            "Value": [
                title,
                school_name,
                start_date,
                end_date,
                strength,
                present,
                absent,
                percentage,
            ],
        }
    )

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:
        overview.to_excel(
            writer,
            sheet_name="Overview",
            index=False,
        )

        daily.to_excel(
            writer,
            sheet_name="Daily Summary",
            index=False,
        )

        details.to_excel(
            writer,
            sheet_name="Class Detail",
            index=False,
        )

        for worksheet in (
            writer.book.worksheets
        ):
            style_excel_sheet(worksheet)

        writer.book[
            "Overview"
        ]["B9"].number_format = "0.00%"

    output.seek(0)

    return output.getvalue()


# =========================================================
# LOGIN AND SCHOOL REGISTRATION
# =========================================================

def render_login_and_registration():
    page_heading(
        "Multi-School Roznamcha Attendance System",
        (
            "Each school has a separate private "
            "account and separate attendance records."
        ),
        show_badge=True,
    )

    sign_in_tab, register_tab = st.tabs(
        [
            "School Sign In",
            "Register New School",
        ]
    )

    with sign_in_tab:
        left, middle, right = st.columns(
            [1, 1.2, 1]
        )

        with middle:
            with st.form(
                "school_login_form"
            ):
                form_heading(
                    "School sign in"
                )

                username = st.text_input(
                    "School username"
                )

                password = st.text_input(
                    "Password",
                    type="password",
                )

                submitted = (
                    st.form_submit_button(
                        "Sign in",
                        use_container_width=True,
                    )
                )

            if submitted:
                user = authenticate(
                    username,
                    password,
                )

                if user:
                    st.session_state[
                        "school_user"
                    ] = user

                    st.rerun()

                else:
                    st.error(
                        "Invalid username "
                        "or password."
                    )

    with register_tab:
        st.info(
            "The master registration password is "
            "only used to create a new school account. "
            "It does not provide access to any school's "
            "attendance."
        )

        with st.form(
            "register_school_form"
        ):
            master = st.text_input(
                "Master registration password",
                type="password",
            )

            school_name = st.text_input(
                "School name"
            )

            username = st.text_input(
                "Create school username"
            )

            first, second = st.columns(2)

            password = first.text_input(
                "Create password",
                type="password",
            )

            confirm = second.text_input(
                "Confirm password",
                type="password",
            )

            register = (
                st.form_submit_button(
                    "Create school account",
                    type="primary",
                )
            )

        if register:
            successful, message = (
                create_school_account(
                    master,
                    school_name,
                    username,
                    password,
                    confirm,
                )
            )

            if successful:
                st.success(message)
            else:
                st.error(message)


# =========================================================
# SCHOOL DASHBOARD
# =========================================================

def render_dashboard(user):
    heading = (
        '<div class="dashboard-heading">'
        '<div class="dashboard-title">'
        f'{html.escape(user["school_name"])}'
        ' — Daily Roznamcha'
        '</div>'
        '<div class="dashboard-subtitle">'
        'Class-wise boys and girls attendance'
        '</div>'
        '</div>'
    )

    st.markdown(
        heading,
        unsafe_allow_html=True,
    )

    selected_date = st.date_input(
        "Attendance date",
        value=date.today(),
    )

    data = attendance_dataframe(
        user["school_id"],
        selected_date,
        selected_date,
    )

    classes = get_classes(
        user["school_id"]
    )

    entered = (
        data["Class"].nunique()
        if not data.empty
        else 0
    )

    if (
        classes
        and entered == len(classes)
    ):
        st.success(
            f"Attendance is complete for "
            f"{selected_date:%d %B %Y}."
        )

    elif entered:
        st.warning(
            f"Attendance is incomplete: "
            f"{entered} of {len(classes)} "
            f"classes entered."
        )

    render_roznamcha_table(data)


# =========================================================
# DAILY ATTENDANCE ENTRY
# =========================================================

def render_attendance_entry(user):
    page_heading(
        "Enter Daily Attendance",
        (
            f"Attendance entry for "
            f"{user['school_name']} only."
        ),
    )

    selected_date = st.date_input(
        "Date",
        value=date.today(),
        key="entry_date",
    )

    classes = get_classes(
        user["school_id"]
    )

    if not classes:
        st.warning(
            "No active classes are configured. "
            "Add a class first."
        )
        return

    existing = get_existing_attendance(
        user["school_id"],
        selected_date,
    )

    entries = []

    st.caption(
        "Absentees and percentages are "
        "calculated automatically."
    )

    form_key = (
        f"attendance_"
        f"{user['school_id']}_"
        f"{selected_date.isoformat()}"
    )

    with st.form(form_key):
        for class_record in classes:
            class_id = class_record["id"]

            saved = existing.get(
                class_id,
                {},
            )

            form_heading(
                f"Class "
                f"{class_record['class_name']}"
            )

            columns = st.columns(4)

            boys_strength = (
                int(
                    class_record[
                        "boys_strength"
                    ]
                )
                if class_record[
                    "boys_enabled"
                ]
                else 0
            )

            girls_strength = (
                int(
                    class_record[
                        "girls_strength"
                    ]
                )
                if class_record[
                    "girls_enabled"
                ]
                else 0
            )

            columns[0].number_input(
                "Boys strength",
                value=boys_strength,
                disabled=True,
                key=(
                    f"bs_"
                    f"{user['school_id']}_"
                    f"{class_id}_"
                    f"{selected_date}"
                ),
            )

            if class_record[
                "boys_enabled"
            ]:
                boys_present = (
                    columns[1].number_input(
                        "Boys present",
                        min_value=0,
                        max_value=boys_strength,
                        value=min(
                            int(
                                saved.get(
                                    "boys_present",
                                    0,
                                )
                            ),
                            boys_strength,
                        ),
                        step=1,
                        key=(
                            f"bp_"
                            f"{user['school_id']}_"
                            f"{class_id}_"
                            f"{selected_date}"
                        ),
                    )
                )

            else:
                columns[1].text_input(
                    "Boys present",
                    value="Not applicable",
                    disabled=True,
                    key=(
                        f"bp_na_"
                        f"{user['school_id']}_"
                        f"{class_id}_"
                        f"{selected_date}"
                    ),
                )

                boys_present = 0

            columns[2].number_input(
                "Girls strength",
                value=girls_strength,
                disabled=True,
                key=(
                    f"gs_"
                    f"{user['school_id']}_"
                    f"{class_id}_"
                    f"{selected_date}"
                ),
            )

            if class_record[
                "girls_enabled"
            ]:
                girls_present = (
                    columns[3].number_input(
                        "Girls present",
                        min_value=0,
                        max_value=girls_strength,
                        value=min(
                            int(
                                saved.get(
                                    "girls_present",
                                    0,
                                )
                            ),
                            girls_strength,
                        ),
                        step=1,
                        key=(
                            f"gp_"
                            f"{user['school_id']}_"
                            f"{class_id}_"
                            f"{selected_date}"
                        ),
                    )
                )

            else:
                columns[3].text_input(
                    "Girls present",
                    value="Not applicable",
                    disabled=True,
                    key=(
                        f"gp_na_"
                        f"{user['school_id']}_"
                        f"{class_id}_"
                        f"{selected_date}"
                    ),
                )

                girls_present = 0

            entries.append(
                {
                    "class_id": class_id,
                    "boys_strength":
                        boys_strength,
                    "girls_strength":
                        girls_strength,
                    "boys_present":
                        int(boys_present),
                    "girls_present":
                        int(girls_present),
                }
            )

            st.divider()

        submitted = (
            st.form_submit_button(
                "Save attendance",
                type="primary",
            )
        )

    if submitted:
        try:
            save_attendance(
                user["school_id"],
                selected_date,
                entries,
            )

            st.success(
                "Attendance saved successfully."
            )

        except ValueError as error:
            st.error(str(error))


# =========================================================
# REPORTS
# =========================================================

def report_period_controls():
    period = st.radio(
        "Report period",
        [
            "Day",
            "Week",
            "Month",
        ],
        horizontal=True,
    )

    reference_date = st.date_input(
        "Select date",
        value=date.today(),
        key="report_date",
    )

    if period == "Day":
        return (
            period,
            reference_date,
            reference_date,
        )

    if period == "Week":
        start = (
            reference_date
            - timedelta(
                days=reference_date.weekday()
            )
        )

        return (
            period,
            start,
            start + timedelta(days=6),
        )

    start = reference_date.replace(
        day=1
    )

    next_month = (
        start.replace(day=28)
        + timedelta(days=4)
    ).replace(day=1)

    return (
        period,
        start,
        next_month - timedelta(days=1),
    )


def render_reports(user):
    page_heading(
        "Attendance Reports",
        (
            f"Only {user['school_name']} "
            f"records are included."
        ),
    )

    (
        period,
        start_date,
        end_date,
    ) = report_period_controls()

    data = attendance_dataframe(
        user["school_id"],
        start_date,
        end_date,
    )

    st.caption(
        f"Report period: "
        f"{start_date:%d %b %Y} to "
        f"{end_date:%d %b %Y}"
    )

    if data.empty:
        st.info(
            "No attendance records exist "
            "for the selected period."
        )
        return

    (
        strength,
        present,
        absent,
        percentage,
    ) = attendance_metrics(data)

    columns = st.columns(4)

    columns[0].metric(
        "Attendance opportunities",
        strength,
    )

    columns[1].metric(
        "Present",
        present,
    )

    columns[2].metric(
        "Absent",
        absent,
    )

    columns[3].metric(
        "Attendance",
        f"{percentage:.2%}",
    )

    preview = daily_summary_dataframe(data)

    st.dataframe(
        preview.style.format(
            {
                "Attendance_Percentage":
                    "{:.2%}"
            }
        ),
        use_container_width=True,
        hide_index=True,
    )

    excel_data = make_excel_report(
        data,
        user["school_name"],
        f"{period} Attendance Report",
        start_date,
        end_date,
    )

    filename = (
        f"attendance_"
        f"{start_date.isoformat()}_to_"
        f"{end_date.isoformat()}.xlsx"
    )

    st.download_button(
        "Download Excel report",
        data=excel_data,
        file_name=filename,
        mime=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        type="primary",
    )


# =========================================================
# SCHOOL AND CLASS MANAGEMENT
# =========================================================

def render_school_and_classes(user):
    page_heading(
        "School and Class Management",
        (
            "Your changes affect only "
            "your own school."
        ),
    )

    section_heading("School name")

    with st.form(
        f"school_name_"
        f"{user['school_id']}"
    ):
        school_name = st.text_input(
            "School name",
            value=user["school_name"],
        )

        save_name = (
            st.form_submit_button(
                "Save school name",
                type="primary",
            )
        )

    if save_name:
        successful, message = (
            update_school_name(
                user["school_id"],
                school_name,
            )
        )

        if successful:
            st.session_state[
                "school_user"
            ]["school_name"] = (
                school_name.strip()
            )

            st.success(message)
            st.rerun()

        else:
            st.error(message)

    classes = get_classes(
        user["school_id"],
        active_only=False,
    )

    section_heading(
        "Manage existing class"
    )

    if classes:
        options = {
            (
                f"{row['class_name']} "
                f"("
                f"{'Active' if row['active'] else 'Inactive'}"
                f")"
            ): row
            for row in classes
        }

        selected_label = st.selectbox(
            "Choose class",
            list(options.keys()),
        )

        selected = options[
            selected_label
        ]

        with st.form(
            f"edit_class_"
            f"{user['school_id']}_"
            f"{selected['id']}"
        ):
            class_name = st.text_input(
                "Class name",
                value=selected["class_name"],
            )

            sort_order = st.number_input(
                "Display order",
                min_value=0,
                value=int(
                    selected["sort_order"]
                ),
                step=1,
            )

            first, second = st.columns(2)

            boys_enabled = first.checkbox(
                "Boys section",
                value=bool(
                    selected["boys_enabled"]
                ),
                key=(
                    f"edit_boys_"
                    f"{user['school_id']}_"
                    f"{selected['id']}"
                ),
            )

            boys_strength = (
                first.number_input(
                    "Boys strength",
                    min_value=0,
                    value=int(
                        selected[
                            "boys_strength"
                        ]
                    ),
                    step=1,
                    key=(
                        f"edit_bs_"
                        f"{user['school_id']}_"
                        f"{selected['id']}"
                    ),
                )
            )

            girls_enabled = (
                second.checkbox(
                    "Girls section",
                    value=bool(
                        selected[
                            "girls_enabled"
                        ]
                    ),
                    key=(
                        f"edit_girls_"
                        f"{user['school_id']}_"
                        f"{selected['id']}"
                    ),
                )
            )

            girls_strength = (
                second.number_input(
                    "Girls strength",
                    min_value=0,
                    value=int(
                        selected[
                            "girls_strength"
                        ]
                    ),
                    step=1,
                    key=(
                        f"edit_gs_"
                        f"{user['school_id']}_"
                        f"{selected['id']}"
                    ),
                )
            )

            active = st.checkbox(
                "Class is active",
                value=bool(
                    selected["active"]
                ),
            )

            save_class = (
                st.form_submit_button(
                    "Save class changes",
                    type="primary",
                )
            )

        if save_class:
            successful, message = (
                update_class(
                    user["school_id"],
                    selected["id"],
                    class_name,
                    sort_order,
                    boys_enabled,
                    girls_enabled,
                    boys_strength,
                    girls_strength,
                    active,
                )
            )

            if successful:
                st.success(message)
                st.rerun()
            else:
                st.error(message)

    else:
        st.info(
            "Your school does not have "
            "any classes yet."
        )

    section_heading(
        "Add another class"
    )

    with st.form(
        f"add_class_"
        f"{user['school_id']}"
    ):
        new_name = st.text_input(
            "New class name"
        )

        new_order = st.number_input(
            "Display order",
            min_value=0,
            value=len(classes) + 1,
            step=1,
        )

        first, second = st.columns(2)

        new_boys = first.checkbox(
            "Boys section",
            value=True,
            key=(
                f"new_boys_"
                f"{user['school_id']}"
            ),
        )

        new_boys_strength = (
            first.number_input(
                "Boys strength",
                min_value=0,
                value=0,
                step=1,
                key=(
                    f"new_bs_"
                    f"{user['school_id']}"
                ),
            )
        )

        new_girls = second.checkbox(
            "Girls section",
            value=True,
            key=(
                f"new_girls_"
                f"{user['school_id']}"
            ),
        )

        new_girls_strength = (
            second.number_input(
                "Girls strength",
                min_value=0,
                value=0,
                step=1,
                key=(
                    f"new_gs_"
                    f"{user['school_id']}"
                ),
            )
        )

        add_button = (
            st.form_submit_button(
                "Add class",
                type="primary",
            )
        )

    if add_button:
        successful, message = add_class(
            user["school_id"],
            new_name,
            new_order,
            new_boys,
            new_girls,
            new_boys_strength,
            new_girls_strength,
        )

        if successful:
            st.success(message)
            st.rerun()
        else:
            st.error(message)


# =========================================================
# ACCOUNT SETTINGS
# =========================================================

def render_account(user):
    page_heading(
        "Login and Password Settings",
        (
            "These credentials belong "
            "only to your school."
        ),
    )

    st.write(
        f"**Current username:** "
        f"{user['username']}"
    )

    section_heading(
        "Change username"
    )

    with st.form(
        f"change_username_"
        f"{user['school_id']}"
    ):
        new_username = st.text_input(
            "New username",
            value=user["username"],
        )

        confirmation_password = (
            st.text_input(
                "Current password to confirm",
                type="password",
            )
        )

        username_button = (
            st.form_submit_button(
                "Change username"
            )
        )

    if username_button:
        successful, message = (
            change_username(
                user["school_id"],
                confirmation_password,
                new_username,
            )
        )

        if successful:
            st.session_state[
                "school_user"
            ]["username"] = (
                new_username.strip()
            )

            st.success(message)
            st.rerun()

        else:
            st.error(message)

    section_heading(
        "Change password"
    )

    with st.form(
        f"change_password_"
        f"{user['school_id']}"
    ):
        current = st.text_input(
            "Current password",
            type="password",
        )

        first, second = st.columns(2)

        new_password = first.text_input(
            "New password",
            type="password",
        )

        confirm = second.text_input(
            "Confirm new password",
            type="password",
        )

        password_button = (
            st.form_submit_button(
                "Change password"
            )
        )

    if password_button:
        if new_password != confirm:
            st.error(
                "The new passwords do not match."
            )

        else:
            successful, message = (
                change_password(
                    user["school_id"],
                    current,
                    new_password,
                )
            )

            if successful:
                st.success(message)
            else:
                st.error(message)


# =========================================================
# PRIVATE SCHOOL NAVIGATION
# =========================================================

def render_application(user):
    fresh_school = get_school(
        user["school_id"]
    )

    if not fresh_school:
        st.session_state.pop(
            "school_user",
            None,
        )

        st.error(
            "This school account "
            "is no longer active."
        )

        st.rerun()

    st.session_state[
        "school_user"
    ].update(
        {
            "school_name":
                fresh_school["school_name"],
            "username":
                fresh_school["username"],
        }
    )

    user = st.session_state[
        "school_user"
    ]

    st.sidebar.markdown(
        '<div class="sidebar-brand">'
        'Roznamcha'
        '</div>',
        unsafe_allow_html=True,
    )

    st.sidebar.caption(
        user["school_name"]
    )

    st.sidebar.caption(
        APP_VERSION
    )

    st.sidebar.write(
        f"Signed in as "
        f"**{user['username']}**"
    )

    selected_page = st.sidebar.radio(
        "Menu",
        [
            "Dashboard",
            "Daily Attendance",
            "Reports",
            "School & Classes",
            "Account",
        ],
    )

    if st.sidebar.button(
        "Sign out",
        use_container_width=True,
    ):
        st.session_state.pop(
            "school_user",
            None,
        )

        st.rerun()

    if selected_page == "Dashboard":
        render_dashboard(user)

    elif selected_page == "Daily Attendance":
        render_attendance_entry(user)

    elif selected_page == "Reports":
        render_reports(user)

    elif selected_page == "School & Classes":
        render_school_and_classes(user)

    else:
        render_account(user)


# =========================================================
# START APPLICATION
# =========================================================

initialize_database()

if "school_user" not in st.session_state:
    render_login_and_registration()

else:
    render_application(
        st.session_state["school_user"]
    )