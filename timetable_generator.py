import streamlit as st
import sqlite3
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DB_NAME = "school_timetable.db"

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]


# ---------------- DATABASE ----------------
def get_connection():
    return sqlite3.connect(DB_NAME, check_same_thread=False)


def run_query(query, params=(), fetch=False):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(query, params)
    data = cur.fetchall()
    conn.commit()
    conn.close()

    if fetch:
        return data


def create_tables():
    run_query("""
        CREATE TABLE IF NOT EXISTS teachers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        )
    """)

    run_query("""
        CREATE TABLE IF NOT EXISTS classes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_name TEXT UNIQUE NOT NULL,
            incharge_teacher TEXT
        )
    """)

    run_query("""
        CREATE TABLE IF NOT EXISTS subjects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_name TEXT UNIQUE NOT NULL
        )
    """)

    run_query("""
        CREATE TABLE IF NOT EXISTS periods (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            period_no INTEGER NOT NULL,
            start_time TEXT,
            end_time TEXT
        )
    """)

    run_query("""
        CREATE TABLE IF NOT EXISTS timetable (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            day TEXT NOT NULL,
            period_no INTEGER NOT NULL,
            class_name TEXT NOT NULL,
            subject_name TEXT NOT NULL,
            teacher_name TEXT NOT NULL
        )
    """)


# ---------------- HELPER FUNCTIONS ----------------
def get_all(table, column):
    data = run_query(f"SELECT {column} FROM {table} ORDER BY {column}", fetch=True)
    return [row[0] for row in data]


def get_periods():
    data = run_query("""
        SELECT period_no, start_time, end_time
        FROM periods
        ORDER BY period_no
    """, fetch=True)
    return data


def get_period_label(period_no):
    data = run_query("""
        SELECT start_time, end_time
        FROM periods
        WHERE period_no=?
        LIMIT 1
    """, (period_no,), fetch=True)

    if data:
        return f"P{period_no}\n{data[0][0]}-{data[0][1]}"
    return f"P{period_no}"


def check_class_conflict(day, period_no, class_name, ignore_id=None):
    """
    One class cannot have two different periods at the same time.
    Teacher conflict is intentionally not checked.
    This allows one teacher to teach more than one class at the same time.
    """

    if ignore_id:
        data = run_query("""
            SELECT *
            FROM timetable
            WHERE day=? AND period_no=? AND class_name=? AND id != ?
        """, (day, period_no, class_name, ignore_id), fetch=True)
    else:
        data = run_query("""
            SELECT *
            FROM timetable
            WHERE day=? AND period_no=? AND class_name=?
        """, (day, period_no, class_name), fetch=True)

    return len(data) > 0


def timetable_dataframe():
    data = run_query("""
        SELECT id, day, period_no, class_name, subject_name, teacher_name
        FROM timetable
        ORDER BY
        CASE day
            WHEN 'Monday' THEN 1
            WHEN 'Tuesday' THEN 2
            WHEN 'Wednesday' THEN 3
            WHEN 'Thursday' THEN 4
            WHEN 'Friday' THEN 5
            WHEN 'Saturday' THEN 6
        END,
        period_no,
        class_name
    """, fetch=True)

    return pd.DataFrame(
        data,
        columns=["ID", "Day", "Period", "Class", "Subject", "Teacher"]
    )


def make_teacher_summary(df):
    if df.empty:
        return pd.DataFrame()

    temp = df.copy()
    temp["Slot"] = temp["Day"] + " - P" + temp["Period"].astype(str)
    temp["Detail"] = temp["Class"] + " / " + temp["Subject"]

    summary = temp.pivot_table(
        index="Teacher",
        columns="Slot",
        values="Detail",
        aggfunc=lambda x: " | ".join(x)
    )

    summary = summary.fillna("")
    summary.reset_index(inplace=True)
    return summary


def make_class_summary(df):
    if df.empty:
        return pd.DataFrame()

    temp = df.copy()
    temp["Slot"] = temp["Day"] + " - P" + temp["Period"].astype(str)
    temp["Detail"] = temp["Subject"] + " / " + temp["Teacher"]

    summary = temp.pivot_table(
        index="Class",
        columns="Slot",
        values="Detail",
        aggfunc=lambda x: " | ".join(x)
    )

    summary = summary.fillna("")
    summary.reset_index(inplace=True)
    return summary


def format_excel_sheet(writer, sheet_name):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.border = border

            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill

    for col in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                value_length = len(str(cell.value))
                if value_length > max_length:
                    max_length = value_length
            except:
                pass

        worksheet.column_dimensions[col_letter].width = min(max_length + 5, 25)

    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True


def make_excel_file():
    output = BytesIO()

    all_df = timetable_dataframe()

    teacher_summary = make_teacher_summary(all_df)
    class_summary = make_class_summary(all_df)

    teachers = get_all("teachers", "name")
    classes = get_all("classes", "class_name")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        all_df.to_excel(writer, sheet_name="Full Timetable", index=False)
        format_excel_sheet(writer, "Full Timetable")

        teacher_summary.to_excel(writer, sheet_name="One Page Teachers", index=False)
        format_excel_sheet(writer, "One Page Teachers")

        class_summary.to_excel(writer, sheet_name="One Page Classes", index=False)
        format_excel_sheet(writer, "One Page Classes")

        for class_name in classes:
            class_df = all_df[all_df["Class"] == class_name]

            if not class_df.empty:
                temp = class_df.copy()
                temp["Period Detail"] = temp["Subject"] + "\n" + temp["Teacher"]

                pivot = temp.pivot_table(
                    index="Day",
                    columns="Period",
                    values="Period Detail",
                    aggfunc=lambda x: " | ".join(x)
                )

                pivot = pivot.reindex(DAYS)
                pivot = pivot.fillna("")

                sheet_name = f"Class {class_name}"[:31]
                pivot.to_excel(writer, sheet_name=sheet_name)
                format_excel_sheet(writer, sheet_name)

        for teacher in teachers:
            teacher_df = all_df[all_df["Teacher"] == teacher]

            if not teacher_df.empty:
                temp = teacher_df.copy()
                temp["Period Detail"] = temp["Class"] + "\n" + temp["Subject"]

                pivot = temp.pivot_table(
                    index="Day",
                    columns="Period",
                    values="Period Detail",
                    aggfunc=lambda x: " | ".join(x)
                )

                pivot = pivot.reindex(DAYS)
                pivot = pivot.fillna("")

                sheet_name = f"Teacher {teacher}"[:31]
                pivot.to_excel(writer, sheet_name=sheet_name)
                format_excel_sheet(writer, sheet_name)

    output.seek(0)
    return output


# ---------------- STREAMLIT APP ----------------
st.set_page_config(
    page_title="School Timetable Generator",
    page_icon="📚",
    layout="wide"
)

create_tables()

st.title("📚 School Timetable Generator")
st.write("Create teacher-wise and class-wise school timetables and download them in Excel format.")


# ---------------- SIDEBAR ----------------
menu = st.sidebar.radio(
    "Select Option",
    [
        "Dashboard",
        "Manage Teachers",
        "Manage Classes",
        "Manage Subjects",
        "Manage Periods",
        "Create Timetable",
        "Edit / Delete Timetable",
        "View Timetable",
        "Download Excel"
    ]
)


# ---------------- DASHBOARD ----------------
if menu == "Dashboard":
    st.header("📊 Dashboard")

    total_teachers = len(get_all("teachers", "name"))
    total_classes = len(get_all("classes", "class_name"))
    total_subjects = len(get_all("subjects", "subject_name"))
    total_periods = len(get_periods())
    total_entries = len(timetable_dataframe())

    col1, col2, col3, col4, col5 = st.columns(5)

    col1.metric("Teachers", total_teachers)
    col2.metric("Classes", total_classes)
    col3.metric("Subjects", total_subjects)
    col4.metric("Periods", total_periods)
    col5.metric("Timetable Entries", total_entries)

    st.info(
        "This updated version allows one teacher to teach more than one class "
        "or subject in the same period."
    )


# ---------------- MANAGE TEACHERS ----------------
elif menu == "Manage Teachers":
    st.header("👨‍🏫 Manage Teachers")

    teacher_name = st.text_input("Enter Teacher Name")

    if st.button("Add Teacher"):
        if teacher_name.strip():
            try:
                run_query(
                    "INSERT INTO teachers (name) VALUES (?)",
                    (teacher_name.strip(),)
                )
                st.success("Teacher added successfully.")
                st.rerun()
            except:
                st.error("This teacher already exists.")
        else:
            st.warning("Please enter teacher name.")

    st.subheader("Existing Teachers")

    teachers = get_all("teachers", "name")

    for teacher in teachers:
        col1, col2 = st.columns([4, 1])
        col1.write(teacher)

        if col2.button("Delete", key=f"delete_teacher_{teacher}"):
            run_query("DELETE FROM teachers WHERE name=?", (teacher,))
            run_query("DELETE FROM timetable WHERE teacher_name=?", (teacher,))
            st.success("Teacher deleted successfully.")
            st.rerun()


# ---------------- MANAGE CLASSES ----------------
elif menu == "Manage Classes":
    st.header("🏫 Manage Classes and Class In-charge")

    teachers = get_all("teachers", "name")

    class_name = st.text_input("Enter Class Name")
    incharge = st.selectbox("Select Class In-charge", [""] + teachers)

    if st.button("Add Class"):
        if class_name.strip():
            try:
                run_query(
                    "INSERT INTO classes (class_name, incharge_teacher) VALUES (?, ?)",
                    (class_name.strip(), incharge)
                )
                st.success("Class added successfully.")
                st.rerun()
            except:
                st.error("This class already exists.")
        else:
            st.warning("Please enter class name.")

    st.subheader("Existing Classes")

    data = run_query("""
        SELECT id, class_name, incharge_teacher
        FROM classes
        ORDER BY class_name
    """, fetch=True)

    for row in data:
        class_id, old_class_name, old_incharge = row

        with st.expander(
            f"{old_class_name} | In-charge: {old_incharge if old_incharge else 'Not Assigned'}"
        ):
            new_class_name = st.text_input(
                "Class Name",
                old_class_name,
                key=f"class_name_{class_id}"
            )

            new_incharge = st.selectbox(
                "Class In-charge",
                [""] + teachers,
                index=([""] + teachers).index(old_incharge)
                if old_incharge in teachers else 0,
                key=f"incharge_{class_id}"
            )

            col1, col2 = st.columns(2)

            if col1.button("Update Class", key=f"update_class_{class_id}"):
                run_query("""
                    UPDATE classes
                    SET class_name=?, incharge_teacher=?
                    WHERE id=?
                """, (new_class_name.strip(), new_incharge, class_id))

                run_query("""
                    UPDATE timetable
                    SET class_name=?
                    WHERE class_name=?
                """, (new_class_name.strip(), old_class_name))

                st.success("Class updated successfully.")
                st.rerun()

            if col2.button("Delete Class", key=f"delete_class_{class_id}"):
                run_query("DELETE FROM classes WHERE id=?", (class_id,))
                run_query("DELETE FROM timetable WHERE class_name=?", (old_class_name,))
                st.success("Class deleted successfully.")
                st.rerun()


# ---------------- MANAGE SUBJECTS ----------------
elif menu == "Manage Subjects":
    st.header("📖 Manage Subjects")

    subject_name = st.text_input("Enter Subject Name")

    if st.button("Add Subject"):
        if subject_name.strip():
            try:
                run_query(
                    "INSERT INTO subjects (subject_name) VALUES (?)",
                    (subject_name.strip(),)
                )
                st.success("Subject added successfully.")
                st.rerun()
            except:
                st.error("This subject already exists.")
        else:
            st.warning("Please enter subject name.")

    st.subheader("Existing Subjects")

    subjects = get_all("subjects", "subject_name")

    for subject in subjects:
        col1, col2 = st.columns([4, 1])
        col1.write(subject)

        if col2.button("Delete", key=f"delete_subject_{subject}"):
            run_query("DELETE FROM subjects WHERE subject_name=?", (subject,))
            run_query("DELETE FROM timetable WHERE subject_name=?", (subject,))
            st.success("Subject deleted successfully.")
            st.rerun()


# ---------------- MANAGE PERIODS ----------------
elif menu == "Manage Periods":
    st.header("⏰ Manage Periods")

    col1, col2, col3 = st.columns(3)

    with col1:
        period_no = st.number_input("Period Number", min_value=1, step=1)

    with col2:
        start_time = st.text_input("Start Time", placeholder="08:00 AM")

    with col3:
        end_time = st.text_input("End Time", placeholder="08:40 AM")

    if st.button("Add Period"):
        if start_time.strip() and end_time.strip():
            run_query("""
                INSERT INTO periods (period_no, start_time, end_time)
                VALUES (?, ?, ?)
            """, (period_no, start_time.strip(), end_time.strip()))

            st.success("Period added successfully.")
            st.rerun()
        else:
            st.warning("Please enter start and end time.")

    st.subheader("Existing Periods")

    periods = get_periods()

    for p in periods:
        p_no, s_time, e_time = p

        with st.expander(f"Period {p_no}: {s_time} - {e_time}"):
            new_p_no = st.number_input(
                "Period No",
                min_value=1,
                value=p_no,
                key=f"period_no_{p_no}_{s_time}"
            )

            new_start = st.text_input(
                "Start Time",
                s_time,
                key=f"start_{p_no}_{s_time}"
            )

            new_end = st.text_input(
                "End Time",
                e_time,
                key=f"end_{p_no}_{s_time}"
            )

            col1, col2 = st.columns(2)

            if col1.button("Update Period", key=f"update_period_{p_no}_{s_time}"):
                run_query("""
                    UPDATE periods
                    SET period_no=?, start_time=?, end_time=?
                    WHERE period_no=? AND start_time=? AND end_time=?
                """, (new_p_no, new_start, new_end, p_no, s_time, e_time))

                run_query("""
                    UPDATE timetable
                    SET period_no=?
                    WHERE period_no=?
                """, (new_p_no, p_no))

                st.success("Period updated successfully.")
                st.rerun()

            if col2.button("Delete Period", key=f"delete_period_{p_no}_{s_time}"):
                run_query("""
                    DELETE FROM periods
                    WHERE period_no=? AND start_time=? AND end_time=?
                """, (p_no, s_time, e_time))

                run_query("DELETE FROM timetable WHERE period_no=?", (p_no,))

                st.success("Period deleted successfully.")
                st.rerun()


# ---------------- CREATE TIMETABLE ----------------
elif menu == "Create Timetable":
    st.header("📝 Create Timetable Entry")

    teachers = get_all("teachers", "name")
    classes = get_all("classes", "class_name")
    subjects = get_all("subjects", "subject_name")
    periods = get_periods()

    if not teachers or not classes or not subjects or not periods:
        st.warning("Please add teachers, classes, subjects, and periods first.")
    else:
        day = st.selectbox("Select Day", DAYS)

        period_no = st.selectbox(
            "Select Period",
            [p[0] for p in periods],
            format_func=lambda x: f"Period {x}"
        )

        class_name = st.selectbox("Select Class", classes)
        subject_name = st.selectbox("Select Subject", subjects)
        teacher_name = st.selectbox("Select Teacher", teachers)

        st.info(
            "Note: The same teacher can be allotted more than one class "
            "or subject in the same period."
        )

        if st.button("Add to Timetable"):

            class_conflict = check_class_conflict(day, period_no, class_name)

            if class_conflict:
                st.error("This class already has a subject in this day and period.")
            else:
                run_query("""
                    INSERT INTO timetable
                    (day, period_no, class_name, subject_name, teacher_name)
                    VALUES (?, ?, ?, ?, ?)
                """, (day, period_no, class_name, subject_name, teacher_name))

                st.success("Timetable entry added successfully.")
                st.rerun()


# ---------------- EDIT / DELETE TIMETABLE ----------------
elif menu == "Edit / Delete Timetable":
    st.header("✏️ Edit or Delete Timetable")

    df = timetable_dataframe()

    if df.empty:
        st.info("No timetable entries available.")
    else:
        st.dataframe(df, use_container_width=True)

        selected_id = st.selectbox("Select Entry ID", df["ID"].tolist())

        selected_row = df[df["ID"] == selected_id].iloc[0]

        teachers = get_all("teachers", "name")
        classes = get_all("classes", "class_name")
        subjects = get_all("subjects", "subject_name")
        periods = get_periods()
        period_numbers = [p[0] for p in periods]

        st.subheader("Edit Selected Entry")

        new_day = st.selectbox(
            "Day",
            DAYS,
            index=DAYS.index(selected_row["Day"])
        )

        new_period = st.selectbox(
            "Period",
            period_numbers,
            index=period_numbers.index(selected_row["Period"])
        )

        new_class = st.selectbox(
            "Class",
            classes,
            index=classes.index(selected_row["Class"])
        )

        new_subject = st.selectbox(
            "Subject",
            subjects,
            index=subjects.index(selected_row["Subject"])
        )

        new_teacher = st.selectbox(
            "Teacher",
            teachers,
            index=teachers.index(selected_row["Teacher"])
        )

        col1, col2 = st.columns(2)

        if col1.button("Update Timetable Entry"):

            class_conflict = check_class_conflict(
                new_day,
                new_period,
                new_class,
                ignore_id=selected_id
            )

            if class_conflict:
                st.error("This class already has another subject in this period.")
            else:
                run_query("""
                    UPDATE timetable
                    SET day=?, period_no=?, class_name=?, subject_name=?, teacher_name=?
                    WHERE id=?
                """, (
                    new_day,
                    new_period,
                    new_class,
                    new_subject,
                    new_teacher,
                    selected_id
                ))

                st.success("Timetable entry updated successfully.")
                st.rerun()

        if col2.button("Delete Timetable Entry"):
            run_query("DELETE FROM timetable WHERE id=?", (selected_id,))
            st.success("Timetable entry deleted successfully.")
            st.rerun()


# ---------------- VIEW TIMETABLE ----------------
elif menu == "View Timetable":
    st.header("👀 View Timetable")

    df = timetable_dataframe()

    if df.empty:
        st.info("No timetable created yet.")
    else:
        view_type = st.radio(
            "View Timetable By",
            [
                "Full Timetable",
                "One Page Teachers Summary",
                "One Page Classes Summary",
                "Class Wise",
                "Teacher Wise"
            ]
        )

        if view_type == "Full Timetable":
            st.dataframe(df, use_container_width=True)

        elif view_type == "One Page Teachers Summary":
            teacher_summary = make_teacher_summary(df)
            st.subheader("One Page Teachers Timetable Summary")
            st.dataframe(teacher_summary, use_container_width=True)

        elif view_type == "One Page Classes Summary":
            class_summary = make_class_summary(df)
            st.subheader("One Page Classes Timetable Summary")
            st.dataframe(class_summary, use_container_width=True)

        elif view_type == "Class Wise":
            classes = get_all("classes", "class_name")
            selected_class = st.selectbox("Select Class", classes)

            class_df = df[df["Class"] == selected_class]

            if class_df.empty:
                st.warning("No timetable found for this class.")
            else:
                st.subheader(f"Timetable for Class {selected_class}")
                st.dataframe(class_df, use_container_width=True)

                temp = class_df.copy()
                temp["Detail"] = temp["Subject"] + " / " + temp["Teacher"]

                pivot = temp.pivot_table(
                    index="Day",
                    columns="Period",
                    values="Detail",
                    aggfunc=lambda x: " | ".join(x)
                )

                pivot = pivot.reindex(DAYS)
                pivot = pivot.fillna("")

                st.subheader("Class Timetable Format")
                st.dataframe(pivot, use_container_width=True)

        elif view_type == "Teacher Wise":
            teachers = get_all("teachers", "name")
            selected_teacher = st.selectbox("Select Teacher", teachers)

            teacher_df = df[df["Teacher"] == selected_teacher]

            if teacher_df.empty:
                st.warning("No timetable found for this teacher.")
            else:
                st.subheader(f"Timetable for Teacher {selected_teacher}")
                st.dataframe(teacher_df, use_container_width=True)

                temp = teacher_df.copy()
                temp["Detail"] = temp["Class"] + " / " + temp["Subject"]

                pivot = temp.pivot_table(
                    index="Day",
                    columns="Period",
                    values="Detail",
                    aggfunc=lambda x: " | ".join(x)
                )

                pivot = pivot.reindex(DAYS)
                pivot = pivot.fillna("")

                st.subheader("Teacher Timetable Format")
                st.dataframe(pivot, use_container_width=True)


# ---------------- DOWNLOAD EXCEL ----------------
elif menu == "Download Excel":
    st.header("⬇️ Download Timetable in Excel")

    df = timetable_dataframe()

    if df.empty:
        st.warning("No timetable available to download.")
    else:
        st.subheader("Full Timetable Preview")
        st.dataframe(df, use_container_width=True)

        st.subheader("One Page Teachers Summary Preview")
        st.dataframe(make_teacher_summary(df), use_container_width=True)

        st.subheader("One Page Classes Summary Preview")
        st.dataframe(make_class_summary(df), use_container_width=True)

        excel_file = make_excel_file()

        file_name = f"school_timetable_{datetime.now().strftime('%Y_%m_%d_%H_%M')}.xlsx"

        st.download_button(
            label="Download Complete Excel Timetable",
            data=excel_file,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )