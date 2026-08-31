import hashlib
import hmac
import html
import os
import secrets
import sqlite3
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill


# =========================================================
# APPLICATION CONFIGURATION
# =========================================================

APP_TITLE = "Markaz School Attendance System"
DB_PATH = os.getenv("ROZNAMCHA_DB", "roznamcha.db")
DEFAULT_CLASSES = ["Katchi", "1", "2", "3", "4", "5"]

st.set_page_config(
    page_title=APP_TITLE,
    page_icon="📘",
    layout="wide"
)


# =========================================================
# DASHBOARD CSS
# st.html() prevents HTML from appearing as text.
# =========================================================

st.html(
    """
    <style>
    .block-container {
        padding-top: 1.5rem;
        max-width: 1600px;
    }

    .dashboard-heading {
        background: linear-gradient(90deg, #071426, #102b49);
        border: 1px solid #274766;
        border-radius: 12px;
        color: white;
        margin-bottom: 0.9rem;
        padding: 0.9rem 1.2rem;
    }

    .dashboard-heading h2 {
        margin: 0;
        color: white;
    }

    .dashboard-heading p {
        margin: 0.25rem 0 0;
        color: #b9d8f4;
    }

    .rozn-scroll {
        overflow-x: auto;
        border-radius: 8px;
        margin: 0.4rem 0 1rem;
    }

    table.rozn-table {
        background: #030303;
        border-collapse: collapse;
        color: #ffffff;
        min-width: 980px;
        text-align: center;
        width: 100%;
    }

    .rozn-table th,
    .rozn-table td {
        border: 1px solid #aeb4bb;
        font-size: 1rem;
        padding: 0.48rem 0.55rem;
        white-space: nowrap;
    }

    .rozn-table thead th {
        font-size: 1.05rem;
        font-weight: 800;
    }

    .rozn-table .class-head,
    .rozn-table .strength-head,
    .rozn-table .strength-cell {
        background: #050505;
    }

    .rozn-table .class-cell {
        background: #050505;
        text-align: left;
    }

    .rozn-table .present-head {
        background: #00df20;
        color: #061b08;
    }

    .rozn-table .present-cell {
        background: #00c921;
        color: #ffffff;
    }

    .rozn-table .absent-head {
        background: #ec1010;
        color: #ffffff;
    }

    .rozn-table .absent-cell {
        background: #671b1b;
        color: #ffffff;
    }

    .rozn-table .percentage-head {
        background: #050505;
        color: #2fe9ff;
    }

    .rozn-table .percentage-cell {
        background: #1e5fae;
        color: #ffffff;
    }

    .rozn-table .total-row td {
        font-weight: 800;
        font-size: 1.06rem;
    }

    table.summary-table {
        background: #061006;
        border-collapse: collapse;
        color: white;
        min-width: 650px;
        text-align: center;
        width: 100%;
    }

    .summary-table th,
    .summary-table td {
        border: 1px solid #aeb4bb;
        padding: 0.48rem 0.5rem;
        white-space: nowrap;
    }

    .summary-table th {
        background: #111820;
        font-weight: 800;
    }

    .summary-table .school-name {
        text-align: left;
    }

    .summary-table .complete {
        color: #5cff79;
        font-weight: 700;
    }

    .summary-table .pending {
        color: #ff7373;
        font-weight: 700;
    }

    .summary-table .summary-total td {
        background: #e000c6;
        color: white;
        font-weight: 800;
    }

    .summary-table .summary-total td:last-child {
        background: #101010;
    }
    </style>
    """
)


# =========================================================
# DATABASE CONNECTION
# =========================================================

@contextmanager
def db_connection():
    connection = sqlite3.connect(
        DB_PATH,
        timeout=30
    )

    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    connection.execute("PRAGMA journal_mode = WAL")

    try:
        yield connection
        connection.commit()

    except Exception:
        connection.rollback()
        raise

    finally:
        connection.close()


# =========================================================
# PASSWORD FUNCTIONS
# =========================================================

def hash_password(password):
    salt = secrets.token_bytes(16)

    digest = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt,
        200_000
    )

    return f"{salt.hex()}${digest.hex()}"


def verify_password(password, stored_password):
    try:
        salt_hex, stored_digest = stored_password.split(
            "$",
            1
        )

        calculated_digest = hashlib.pbkdf2_hmac(
            "sha256",
            password.encode("utf-8"),
            bytes.fromhex(salt_hex),
            200_000
        ).hex()

        return hmac.compare_digest(
            calculated_digest,
            stored_digest
        )

    except (ValueError, AttributeError):
        return False


# =========================================================
# DATABASE INITIALIZATION
# =========================================================

def initialize_database():
    with db_connection() as con:

        con.executescript(
            """
            CREATE TABLE IF NOT EXISTS schools (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE COLLATE NOCASE,
                password_hash TEXT NOT NULL,
                display_name TEXT NOT NULL,
                role TEXT NOT NULL
                    CHECK (role IN ('AEO', 'SCHOOL')),
                school_id INTEGER,
                active INTEGER NOT NULL DEFAULT 1,
                must_change_password INTEGER NOT NULL DEFAULT 1,
                FOREIGN KEY (school_id) REFERENCES schools(id)
            );

            CREATE TABLE IF NOT EXISTS school_classes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                school_id INTEGER NOT NULL,
                class_name TEXT NOT NULL,
                sort_order INTEGER NOT NULL DEFAULT 0,
                boys_enabled INTEGER NOT NULL DEFAULT 1,
                girls_enabled INTEGER NOT NULL DEFAULT 1,
                boys_strength INTEGER NOT NULL DEFAULT 0
                    CHECK (boys_strength >= 0),
                girls_strength INTEGER NOT NULL DEFAULT 0
                    CHECK (girls_strength >= 0),
                active INTEGER NOT NULL DEFAULT 1,
                UNIQUE (school_id, class_name),
                FOREIGN KEY (school_id) REFERENCES schools(id)
            );

            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                school_id INTEGER NOT NULL,
                class_id INTEGER NOT NULL,
                attendance_date TEXT NOT NULL,
                boys_strength INTEGER NOT NULL DEFAULT 0,
                girls_strength INTEGER NOT NULL DEFAULT 0,
                boys_present INTEGER NOT NULL DEFAULT 0,
                girls_present INTEGER NOT NULL DEFAULT 0,
                entered_by INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE (
                    school_id,
                    class_id,
                    attendance_date
                ),
                FOREIGN KEY (school_id) REFERENCES schools(id),
                FOREIGN KEY (class_id) REFERENCES school_classes(id),
                FOREIGN KEY (entered_by) REFERENCES users(id)
            );

            CREATE INDEX IF NOT EXISTS idx_attendance_date
            ON attendance(attendance_date);

            CREATE INDEX IF NOT EXISTS idx_attendance_school_date
            ON attendance(school_id, attendance_date);
            """
        )

        school_count = con.execute(
            "SELECT COUNT(*) FROM schools"
        ).fetchone()[0]

        if school_count == 0:

            for number in range(1, 15):
                school_code = f"SCH{number:02d}"
                school_name = f"School {number:02d}"

                cursor = con.execute(
                    """
                    INSERT INTO schools(code, name)
                    VALUES (?, ?)
                    """,
                    (
                        school_code,
                        school_name
                    )
                )

                school_id = cursor.lastrowid

                for order, class_name in enumerate(
                    DEFAULT_CLASSES,
                    start=1
                ):
                    con.execute(
                        """
                        INSERT INTO school_classes(
                            school_id,
                            class_name,
                            sort_order,
                            boys_enabled,
                            girls_enabled,
                            boys_strength,
                            girls_strength
                        )
                        VALUES (?, ?, ?, 1, 1, 0, 0)
                        """,
                        (
                            school_id,
                            class_name,
                            order
                        )
                    )

                con.execute(
                    """
                    INSERT INTO users(
                        username,
                        password_hash,
                        display_name,
                        role,
                        school_id
                    )
                    VALUES (?, ?, ?, 'SCHOOL', ?)
                    """,
                    (
                        f"school{number:02d}",
                        hash_password(
                            f"School{number:02d}@2026!"
                        ),
                        f"{school_name} User",
                        school_id
                    )
                )

        aeo_exists = con.execute(
            """
            SELECT 1
            FROM users
            WHERE role = 'AEO'
            LIMIT 1
            """
        ).fetchone()

        if not aeo_exists:
            con.execute(
                """
                INSERT INTO users(
                    username,
                    password_hash,
                    display_name,
                    role,
                    school_id,
                    must_change_password
                )
                VALUES (?, ?, ?, 'AEO', NULL, 1)
                """,
                (
                    "aeo",
                    hash_password("AEO@2026!"),
                    "AEO Markaz"
                )
            )


# =========================================================
# AUTHENTICATION
# =========================================================

def authenticate(username, password):
    with db_connection() as con:

        user = con.execute(
            """
            SELECT
                u.*,
                s.name AS school_name
            FROM users u
            LEFT JOIN schools s
                ON s.id = u.school_id
            WHERE u.username = ?
              AND u.active = 1
              AND (
                    u.school_id IS NULL
                    OR s.active = 1
                  )
            """,
            (username.strip(),)
        ).fetchone()

    if user and verify_password(
        password,
        user["password_hash"]
    ):
        return dict(user)

    return None


def change_password(
    user_id,
    current_password,
    new_password
):
    with db_connection() as con:

        user = con.execute(
            """
            SELECT password_hash
            FROM users
            WHERE id = ?
            """,
            (user_id,)
        ).fetchone()

        if not user:
            return False, "User account was not found."

        if not verify_password(
            current_password,
            user["password_hash"]
        ):
            return False, "The current password is incorrect."

        if len(new_password) < 8:
            return (
                False,
                "The new password must contain at least 8 characters."
            )

        con.execute(
            """
            UPDATE users
            SET
                password_hash = ?,
                must_change_password = 0
            WHERE id = ?
            """,
            (
                hash_password(new_password),
                user_id
            )
        )

    return True, "Password changed successfully."


def change_own_username(
    user_id,
    current_password,
    new_username
):
    new_username = new_username.strip()

    if len(new_username) < 3:
        return (
            False,
            "The username must contain at least 3 characters."
        )

    if " " in new_username:
        return (
            False,
            "The username cannot contain spaces."
        )

    try:
        with db_connection() as con:

            user = con.execute(
                """
                SELECT password_hash
                FROM users
                WHERE id = ?
                """,
                (user_id,)
            ).fetchone()

            if not user or not verify_password(
                current_password,
                user["password_hash"]
            ):
                return (
                    False,
                    "The current password is incorrect."
                )

            con.execute(
                """
                UPDATE users
                SET username = ?
                WHERE id = ?
                """,
                (
                    new_username,
                    user_id
                )
            )

        return True, "Username changed successfully."

    except sqlite3.IntegrityError:
        return (
            False,
            "This username is already being used."
        )


# =========================================================
# GENERAL DATABASE FUNCTIONS
# =========================================================

def fetch_dataframe(sql, parameters=()):
    with db_connection() as con:
        return pd.read_sql_query(
            sql,
            con,
            params=parameters
        )


def get_active_schools():
    with db_connection() as con:
        return con.execute(
            """
            SELECT id, code, name
            FROM schools
            WHERE active = 1
            ORDER BY code
            """
        ).fetchall()


def get_school_user(school_id):
    with db_connection() as con:
        return con.execute(
            """
            SELECT
                id,
                username,
                display_name
            FROM users
            WHERE school_id = ?
              AND role = 'SCHOOL'
            ORDER BY id
            LIMIT 1
            """,
            (school_id,)
        ).fetchone()


def get_school_classes(
    school_id,
    active_only=True
):
    active_condition = (
        "AND active = 1"
        if active_only
        else ""
    )

    with db_connection() as con:
        return con.execute(
            f"""
            SELECT *
            FROM school_classes
            WHERE school_id = ?
            {active_condition}
            ORDER BY sort_order, class_name
            """,
            (school_id,)
        ).fetchall()


def get_existing_attendance(
    school_id,
    attendance_date
):
    with db_connection() as con:

        records = con.execute(
            """
            SELECT
                class_id,
                boys_present,
                girls_present
            FROM attendance
            WHERE school_id = ?
              AND attendance_date = ?
            """,
            (
                school_id,
                attendance_date.isoformat()
            )
        ).fetchall()

    return {
        record["class_id"]: dict(record)
        for record in records
    }


# =========================================================
# SAVE ATTENDANCE
# =========================================================

def save_attendance(
    school_id,
    selected_date,
    user_id,
    attendance_entries
):
    timestamp = (
        datetime.now()
        .replace(microsecond=0)
        .isoformat()
    )

    with db_connection() as con:

        for entry in attendance_entries:

            con.execute(
                """
                INSERT INTO attendance(
                    school_id,
                    class_id,
                    attendance_date,
                    boys_strength,
                    girls_strength,
                    boys_present,
                    girls_present,
                    entered_by,
                    created_at,
                    updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)

                ON CONFLICT(
                    school_id,
                    class_id,
                    attendance_date
                )

                DO UPDATE SET
                    boys_strength =
                        excluded.boys_strength,
                    girls_strength =
                        excluded.girls_strength,
                    boys_present =
                        excluded.boys_present,
                    girls_present =
                        excluded.girls_present,
                    entered_by =
                        excluded.entered_by,
                    updated_at =
                        excluded.updated_at
                """,
                (
                    school_id,
                    entry["class_id"],
                    selected_date.isoformat(),
                    entry["boys_strength"],
                    entry["girls_strength"],
                    entry["boys_present"],
                    entry["girls_present"],
                    user_id,
                    timestamp,
                    timestamp
                )
            )


# =========================================================
# ATTENDANCE DATA
# =========================================================

ATTENDANCE_QUERY = """
SELECT
    a.attendance_date AS Date,
    s.id AS School_ID,
    s.code AS School_Code,
    s.name AS School,
    c.class_name AS Class,
    c.sort_order AS Class_Order,
    a.boys_strength AS Boys_Strength,
    a.girls_strength AS Girls_Strength,
    a.boys_present AS Boys_Present,
    a.girls_present AS Girls_Present
FROM attendance a
JOIN schools s
    ON s.id = a.school_id
JOIN school_classes c
    ON c.id = a.class_id
WHERE a.attendance_date BETWEEN ? AND ?
"""


def attendance_dataframe(
    start_date,
    end_date,
    school_id=None
):
    sql = ATTENDANCE_QUERY

    parameters = [
        start_date.isoformat(),
        end_date.isoformat()
    ]

    if school_id is not None:
        sql += " AND s.id = ?"
        parameters.append(school_id)

    sql += """
        ORDER BY
            a.attendance_date,
            s.code,
            c.sort_order,
            c.class_name
    """

    data = fetch_dataframe(
        sql,
        parameters
    )

    if data.empty:
        return data

    numeric_columns = [
        "Boys_Strength",
        "Girls_Strength",
        "Boys_Present",
        "Girls_Present"
    ]

    data[numeric_columns] = (
        data[numeric_columns]
        .fillna(0)
        .astype(int)
    )

    data["Total_Strength"] = (
        data["Boys_Strength"]
        + data["Girls_Strength"]
    )

    data["Boys_Absent"] = (
        data["Boys_Strength"]
        - data["Boys_Present"]
    )

    data["Girls_Absent"] = (
        data["Girls_Strength"]
        - data["Girls_Present"]
    )

    data["Total_Present"] = (
        data["Boys_Present"]
        + data["Girls_Present"]
    )

    data["Total_Absent"] = (
        data["Boys_Absent"]
        + data["Girls_Absent"]
    )

    data["Attendance_Percentage"] = (
        data["Total_Present"]
        .div(
            data["Total_Strength"]
            .replace(0, pd.NA)
        )
        .fillna(0)
    )

    data["Date"] = (
        pd.to_datetime(data["Date"])
        .dt.date
    )

    return data


def calculate_metrics(data):
    if data.empty:
        return 0, 0, 0, 0.0

    strength = int(
        data["Total_Strength"].sum()
    )

    present = int(
        data["Total_Present"].sum()
    )

    absent = int(
        data["Total_Absent"].sum()
    )

    percentage = (
        present / strength
        if strength
        else 0.0
    )

    return (
        strength,
        present,
        absent,
        percentage
    )


def summarize_attendance(
    data,
    grouping_columns
):
    if data.empty:
        return pd.DataFrame()

    summary = (
        data
        .groupby(
            grouping_columns,
            as_index=False
        )
        .agg(
            Attendance_Opportunities=(
                "Total_Strength",
                "sum"
            ),
            Total_Present=(
                "Total_Present",
                "sum"
            ),
            Total_Absent=(
                "Total_Absent",
                "sum"
            )
        )
    )

    summary["Attendance_Percentage"] = (
        summary["Total_Present"]
        .div(
            summary[
                "Attendance_Opportunities"
            ].replace(0, pd.NA)
        )
        .fillna(0)
    )

    return summary


# =========================================================
# WORKBOOK-STYLE ATTENDANCE TABLE
# =========================================================

def class_summary_dataframe(data):
    if data.empty:
        return pd.DataFrame()

    summary = (
        data
        .groupby(
            "Class",
            as_index=False
        )
        .agg(
            Class_Order=(
                "Class_Order",
                "min"
            ),
            Boys_Strength=(
                "Boys_Strength",
                "sum"
            ),
            Girls_Strength=(
                "Girls_Strength",
                "sum"
            ),
            Boys_Present=(
                "Boys_Present",
                "sum"
            ),
            Girls_Present=(
                "Girls_Present",
                "sum"
            )
        )
        .sort_values(
            ["Class_Order", "Class"],
            kind="stable"
        )
    )

    summary["Total_Strength"] = (
        summary["Boys_Strength"]
        + summary["Girls_Strength"]
    )

    summary["Total_Present"] = (
        summary["Boys_Present"]
        + summary["Girls_Present"]
    )

    summary["Boys_Absent"] = (
        summary["Boys_Strength"]
        - summary["Boys_Present"]
    )

    summary["Girls_Absent"] = (
        summary["Girls_Strength"]
        - summary["Girls_Present"]
    )

    summary["Total_Absent"] = (
        summary["Boys_Absent"]
        + summary["Girls_Absent"]
    )

    summary["Attendance_Percentage"] = (
        summary["Total_Present"]
        .div(
            summary["Total_Strength"]
            .replace(0, pd.NA)
        )
        .fillna(0)
    )

    return summary


def render_roznamcha_table(data):
    summary = class_summary_dataframe(data)

    if summary.empty:
        st.info(
            "Attendance has not been entered "
            "for this date."
        )
        return

    body_rows = []

    for _, row in summary.iterrows():

        body_rows.append(
            f"""
            <tr>
                <td class="class-cell">
                    {html.escape(str(row["Class"]))}
                </td>
                <td class="strength-cell">
                    {int(row["Boys_Strength"])}
                </td>
                <td class="strength-cell">
                    {int(row["Girls_Strength"])}
                </td>
                <td class="strength-cell">
                    {int(row["Total_Strength"])}
                </td>
                <td class="present-cell">
                    {int(row["Boys_Present"])}
                </td>
                <td class="present-cell">
                    {int(row["Girls_Present"])}
                </td>
                <td class="present-cell">
                    {int(row["Total_Present"])}
                </td>
                <td class="absent-cell">
                    {int(row["Boys_Absent"])}
                </td>
                <td class="absent-cell">
                    {int(row["Girls_Absent"])}
                </td>
                <td class="absent-cell">
                    {int(row["Total_Absent"])}
                </td>
                <td class="percentage-cell">
                    {row["Attendance_Percentage"]:.2%}
                </td>
            </tr>
            """
        )

    total_columns = [
        "Boys_Strength",
        "Girls_Strength",
        "Total_Strength",
        "Boys_Present",
        "Girls_Present",
        "Total_Present",
        "Boys_Absent",
        "Girls_Absent",
        "Total_Absent"
    ]

    totals = {
        column: int(summary[column].sum())
        for column in total_columns
    }

    total_percentage = (
        totals["Total_Present"]
        / totals["Total_Strength"]
        if totals["Total_Strength"]
        else 0.0
    )

    table_html = f"""
    <div class="rozn-scroll">
        <table class="rozn-table">
            <thead>
                <tr>
                    <th class="class-head" rowspan="2">
                        Classes
                    </th>
                    <th class="strength-head" colspan="3">
                        Total Strength
                    </th>
                    <th class="present-head" colspan="3">
                        Present
                    </th>
                    <th class="absent-head" colspan="3">
                        Absentees
                    </th>
                    <th class="percentage-head" rowspan="2">
                        Percentage<br>Attendance
                    </th>
                </tr>
                <tr>
                    <th class="strength-head">Boys</th>
                    <th class="strength-head">Girls</th>
                    <th class="strength-head">Total</th>
                    <th class="present-head">Boys</th>
                    <th class="present-head">Girls</th>
                    <th class="present-head">Total</th>
                    <th class="absent-head">Boys</th>
                    <th class="absent-head">Girls</th>
                    <th class="absent-head">Total</th>
                </tr>
            </thead>
            <tbody>
                {''.join(body_rows)}
                <tr class="total-row">
                    <td class="class-cell">Total</td>
                    <td class="strength-cell">
                        {totals["Boys_Strength"]}
                    </td>
                    <td class="strength-cell">
                        {totals["Girls_Strength"]}
                    </td>
                    <td class="strength-cell">
                        {totals["Total_Strength"]}
                    </td>
                    <td class="present-cell">
                        {totals["Boys_Present"]}
                    </td>
                    <td class="present-cell">
                        {totals["Girls_Present"]}
                    </td>
                    <td class="present-cell">
                        {totals["Total_Present"]}
                    </td>
                    <td class="absent-cell">
                        {totals["Boys_Absent"]}
                    </td>
                    <td class="absent-cell">
                        {totals["Girls_Absent"]}
                    </td>
                    <td class="absent-cell">
                        {totals["Total_Absent"]}
                    </td>
                    <td class="percentage-cell">
                        {total_percentage:.2%}
                    </td>
                </tr>
            </tbody>
        </table>
    </div>
    """

    # Direct HTML rendering fixes the displayed HTML tags.
    st.html(table_html)


# =========================================================
# AEO SCHOOL SUMMARY TABLE
# =========================================================

def render_school_status_table(summary_rows):
    body_rows = []

    for row in summary_rows:

        status_class = (
            "complete"
            if row["Status"] == "Complete"
            else "pending"
        )

        body_rows.append(
            f"""
            <tr>
                <td class="school-name">
                    {html.escape(row["School"])}
                </td>
                <td>{row["Strength"]}</td>
                <td>{row["Present"]}</td>
                <td>{row["Absent"]}</td>
                <td>{row["Attendance %"]:.2%}</td>
                <td class="{status_class}">
                    {html.escape(row["Status"])}
                </td>
            </tr>
            """
        )

    total_strength = sum(
        row["Strength"]
        for row in summary_rows
    )

    total_present = sum(
        row["Present"]
        for row in summary_rows
    )

    total_absent = sum(
        row["Absent"]
        for row in summary_rows
    )

    total_percentage = (
        total_present / total_strength
        if total_strength
        else 0.0
    )

    summary_html = f"""
    <div class="rozn-scroll">
        <table class="summary-table">
            <thead>
                <tr>
                    <th>School</th>
                    <th>Strength</th>
                    <th>Present</th>
                    <th>Absent</th>
                    <th>Percentage</th>
                    <th>Status</th>
                </tr>
            </thead>
            <tbody>
                {''.join(body_rows)}
                <tr class="summary-total">
                    <td>Total</td>
                    <td>{total_strength}</td>
                    <td>{total_present}</td>
                    <td>{total_absent}</td>
                    <td>{total_percentage:.2%}</td>
                    <td>Markaz</td>
                </tr>
            </tbody>
        </table>
    </div>
    """

    st.html(summary_html)


# =========================================================
# EXCEL REPORT
# =========================================================

def style_excel_sheet(worksheet):
    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:

        values = [
            str(cell.value)
            if cell.value is not None
            else ""
            for cell in column_cells
        ]

        maximum_length = max(
            map(len, values),
            default=0
        )

        width = min(
            max(maximum_length + 2, 11),
            30
        )

        worksheet.column_dimensions[
            column_cells[0].column_letter
        ].width = width

    for row in worksheet.iter_rows(min_row=2):

        for cell in row:

            heading = str(
                worksheet.cell(
                    1,
                    cell.column
                ).value
            )

            if "Percentage" in heading:
                cell.number_format = "0.00%"


def make_excel_report(
    data,
    title,
    start_date,
    end_date
):
    output = BytesIO()

    class_detail = data[
        [
            "Date",
            "School_Code",
            "School",
            "Class",
            "Boys_Strength",
            "Girls_Strength",
            "Total_Strength",
            "Boys_Present",
            "Girls_Present",
            "Total_Present",
            "Boys_Absent",
            "Girls_Absent",
            "Total_Absent",
            "Attendance_Percentage"
        ]
    ].copy()

    daily_summary = summarize_attendance(
        data,
        [
            "Date",
            "School_Code",
            "School"
        ]
    )

    school_summary = summarize_attendance(
        data,
        [
            "School_Code",
            "School"
        ]
    )

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        school_summary.to_excel(
            writer,
            sheet_name="School Summary",
            index=False
        )

        daily_summary.to_excel(
            writer,
            sheet_name="Daily Summary",
            index=False
        )

        class_detail.to_excel(
            writer,
            sheet_name="Class Detail",
            index=False
        )

        workbook = writer.book

        for worksheet in workbook.worksheets:
            style_excel_sheet(worksheet)

        summary_sheet = workbook[
            "School Summary"
        ]

        summary_sheet["G1"] = title
        summary_sheet["G2"] = (
            f"Period: {start_date.isoformat()} "
            f"to {end_date.isoformat()}"
        )

        summary_sheet["G1"].font = Font(
            bold=True,
            size=14
        )

    output.seek(0)

    return output.getvalue()


# =========================================================
# LOGIN
# =========================================================

def render_login():
    st.title(
        "📘 Markaz School Attendance System"
    )

    st.caption(
        "Daily class-wise attendance for schools "
        "and the AEO Markaz dashboard"
    )

    left, middle, right = st.columns(
        [1, 1.2, 1]
    )

    with middle:
        with st.form("login_form"):

            st.subheader("Sign in")

            username = st.text_input(
                "Username"
            )

            password = st.text_input(
                "Password",
                type="password"
            )

            submitted = st.form_submit_button(
                "Sign in",
                use_container_width=True
            )

        if submitted:

            user = authenticate(
                username,
                password
            )

            if user:
                st.session_state.user = user
                st.rerun()

            else:
                st.error(
                    "Invalid username or password."
                )


def render_required_password_change(user):
    st.title(
        "Change your temporary password"
    )

    st.warning(
        "You must choose a new password "
        "before using the app."
    )

    with st.form(
        "required_password_change"
    ):
        current_password = st.text_input(
            "Temporary/current password",
            type="password"
        )

        new_password = st.text_input(
            "New password",
            type="password"
        )

        confirm_password = st.text_input(
            "Confirm new password",
            type="password"
        )

        submitted = st.form_submit_button(
            "Change password"
        )

    if submitted:

        if new_password != confirm_password:
            st.error(
                "The new passwords do not match."
            )

        else:
            successful, message = change_password(
                user["id"],
                current_password,
                new_password
            )

            if successful:
                st.session_state.pop(
                    "user",
                    None
                )

                st.success(
                    message
                    + " Please sign in again."
                )

                st.rerun()

            else:
                st.error(message)


# =========================================================
# SCHOOL DASHBOARD
# =========================================================

def render_school_dashboard(user):
    heading_html = f"""
    <div class="dashboard-heading">
        <h2>
            {html.escape(user["school_name"])}
            — Daily Roznamcha
        </h2>
        <p>
            Class-wise boys and girls attendance
        </p>
    </div>
    """

    st.html(heading_html)

    selected_date = st.date_input(
        "Attendance date",
        value=date.today()
    )

    data = attendance_dataframe(
        selected_date,
        selected_date,
        user["school_id"]
    )

    classes = get_school_classes(
        user["school_id"]
    )

    entered_classes = (
        data["Class"].nunique()
        if not data.empty
        else 0
    )

    if (
        entered_classes == len(classes)
        and classes
    ):
        st.success(
            f"Attendance is complete for "
            f"{selected_date:%d %B %Y}."
        )

    elif entered_classes:
        st.warning(
            f"Attendance is incomplete: "
            f"{entered_classes} of "
            f"{len(classes)} classes entered."
        )

    render_roznamcha_table(data)


# =========================================================
# AEO DASHBOARD
# =========================================================

def render_aeo_dashboard():
    heading_html = """
    <div class="dashboard-heading">
        <h2>AEO Markaz Attendance Dashboard</h2>
        <p>
            Markaz totals, individual-school attendance
            and daily submission monitoring
        </p>
    </div>
    """

    st.html(heading_html)

    selected_date = st.date_input(
        "Attendance date",
        value=date.today()
    )

    data = attendance_dataframe(
        selected_date,
        selected_date
    )

    schools = get_active_schools()

    submitted_school_ids = (
        set(data["School_ID"].unique())
        if not data.empty
        else set()
    )

    (
        total_strength,
        total_present,
        total_absent,
        total_percentage
    ) = calculate_metrics(data)

    metric_columns = st.columns(5)

    metric_columns[0].metric(
        "Schools",
        len(schools)
    )

    metric_columns[1].metric(
        "Submitted",
        len(submitted_school_ids)
    )

    metric_columns[2].metric(
        "Reported strength",
        total_strength
    )

    metric_columns[3].metric(
        "Present",
        total_present
    )

    metric_columns[4].metric(
        "Markaz attendance",
        f"{total_percentage:.2%}"
    )

    school_summary_rows = []

    for school in schools:

        if data.empty:
            school_data = data

        else:
            school_data = data[
                data["School_ID"]
                == school["id"]
            ]

        configured_classes = len(
            get_school_classes(
                school["id"]
            )
        )

        entered_classes = (
            len(school_data)
            if not school_data.empty
            else 0
        )

        (
            school_strength,
            school_present,
            school_absent,
            school_percentage
        ) = calculate_metrics(school_data)

        if entered_classes == 0:
            status = "Not submitted"

        elif entered_classes < configured_classes:
            status = (
                f"Incomplete "
                f"({entered_classes}/"
                f"{configured_classes})"
            )

        else:
            status = "Complete"

        school_summary_rows.append(
            {
                "Code": school["code"],
                "School": school["name"],
                "Status": status,
                "Strength": school_strength,
                "Present": school_present,
                "Absent": school_absent,
                "Attendance %": school_percentage
            }
        )

    view_options = {
        "Markaz Total — all submitted schools": None
    }

    view_options.update(
        {
            (
                f"{school['code']} — "
                f"{school['name']}"
            ): school["id"]
            for school in schools
        }
    )

    selected_view = st.selectbox(
        "Attendance table to display",
        list(view_options.keys())
    )

    selected_school_id = view_options[
        selected_view
    ]

    if selected_school_id is None:
        selected_data = data

    else:
        selected_data = data[
            data["School_ID"]
            == selected_school_id
        ]

    st.subheader(selected_view)

    render_roznamcha_table(
        selected_data
    )

    st.subheader(
        "Daily school submission summary"
    )

    render_school_status_table(
        school_summary_rows
    )

    pending_schools = [
        row["School"]
        for row in school_summary_rows
        if row["Status"] != "Complete"
    ]

    if pending_schools:
        st.warning(
            "Pending or incomplete schools: "
            + ", ".join(pending_schools)
        )


# =========================================================
# DAILY ATTENDANCE
# =========================================================

def render_attendance_entry(user):
    st.title("Enter Daily Attendance")

    selected_date = st.date_input(
        "Date",
        value=date.today()
    )

    classes = get_school_classes(
        user["school_id"]
    )

    if not classes:
        st.warning(
            "No active classes are configured. "
            "Add classes first."
        )
        return

    existing_attendance = get_existing_attendance(
        user["school_id"],
        selected_date
    )

    st.caption(
        "Absentees and percentages are "
        "calculated automatically."
    )

    attendance_entries = []

    with st.form(
        f"attendance_{selected_date.isoformat()}"
    ):

        for class_record in classes:

            saved_record = existing_attendance.get(
                class_record["id"],
                {}
            )

            # This is intentional Markdown, not HTML.
            st.markdown(
                f"#### Class "
                f"{class_record['class_name']}"
            )

            columns = st.columns(4)

            boys_strength = (
                class_record["boys_strength"]
                if class_record["boys_enabled"]
                else 0
            )

            girls_strength = (
                class_record["girls_strength"]
                if class_record["girls_enabled"]
                else 0
            )

            columns[0].number_input(
                "Boys strength",
                value=int(boys_strength),
                disabled=True,
                key=(
                    f"boys_strength_"
                    f"{class_record['id']}_"
                    f"{selected_date}"
                )
            )

            if class_record["boys_enabled"]:

                boys_present = columns[1].number_input(
                    "Boys present",
                    min_value=0,
                    max_value=int(boys_strength),
                    value=min(
                        int(
                            saved_record.get(
                                "boys_present",
                                0
                            )
                        ),
                        int(boys_strength)
                    ),
                    step=1,
                    key=(
                        f"boys_present_"
                        f"{class_record['id']}_"
                        f"{selected_date}"
                    )
                )

            else:
                columns[1].text_input(
                    "Boys present",
                    value="Not applicable",
                    disabled=True,
                    key=(
                        f"boys_na_"
                        f"{class_record['id']}_"
                        f"{selected_date}"
                    )
                )

                boys_present = 0

            columns[2].number_input(
                "Girls strength",
                value=int(girls_strength),
                disabled=True,
                key=(
                    f"girls_strength_"
                    f"{class_record['id']}_"
                    f"{selected_date}"
                )
            )

            if class_record["girls_enabled"]:

                girls_present = columns[3].number_input(
                    "Girls present",
                    min_value=0,
                    max_value=int(girls_strength),
                    value=min(
                        int(
                            saved_record.get(
                                "girls_present",
                                0
                            )
                        ),
                        int(girls_strength)
                    ),
                    step=1,
                    key=(
                        f"girls_present_"
                        f"{class_record['id']}_"
                        f"{selected_date}"
                    )
                )

            else:
                columns[3].text_input(
                    "Girls present",
                    value="Not applicable",
                    disabled=True,
                    key=(
                        f"girls_na_"
                        f"{class_record['id']}_"
                        f"{selected_date}"
                    )
                )

                girls_present = 0

            attendance_entries.append(
                {
                    "class_id":
                        class_record["id"],

                    "boys_strength":
                        int(boys_strength),

                    "girls_strength":
                        int(girls_strength),

                    "boys_present":
                        int(boys_present),

                    "girls_present":
                        int(girls_present)
                }
            )

            st.divider()

        submitted = st.form_submit_button(
            "Save attendance",
            type="primary"
        )

    if submitted:
        save_attendance(
            user["school_id"],
            selected_date,
            user["id"],
            attendance_entries
        )

        st.success(
            "Attendance saved successfully. "
            "You can reopen this date to update it."
        )


# =========================================================
# REPORTS
# =========================================================

def report_period_controls():
    period = st.radio(
        "Report period",
        [
            "Day",
            "Week",
            "Month"
        ],
        horizontal=True
    )

    reference_date = st.date_input(
        "Select date",
        value=date.today(),
        key="report_reference_date"
    )

    if period == "Day":
        return (
            period,
            reference_date,
            reference_date
        )

    if period == "Week":
        start_date = (
            reference_date
            - timedelta(
                days=reference_date.weekday()
            )
        )

        end_date = (
            start_date
            + timedelta(days=6)
        )

        return (
            period,
            start_date,
            end_date
        )

    start_date = reference_date.replace(
        day=1
    )

    next_month = (
        start_date.replace(day=28)
        + timedelta(days=4)
    ).replace(day=1)

    end_date = (
        next_month
        - timedelta(days=1)
    )

    return (
        period,
        start_date,
        end_date
    )


def render_reports(user):
    st.title("Attendance Reports")

    (
        period,
        start_date,
        end_date
    ) = report_period_controls()

    school_id = user["school_id"]

    selected_name = (
        user.get("school_name")
        or "All Schools"
    )

    if user["role"] == "AEO":

        schools = get_active_schools()

        school_options = {
            "All Schools (Markaz)": None
        }

        school_options.update(
            {
                (
                    f"{school['code']} — "
                    f"{school['name']}"
                ): school["id"]
                for school in schools
            }
        )

        selected_name = st.selectbox(
            "School",
            list(school_options.keys())
        )

        school_id = school_options[
            selected_name
        ]

    data = attendance_dataframe(
        start_date,
        end_date,
        school_id
    )

    st.caption(
        f"Report period: "
        f"{start_date:%d %b %Y} "
        f"to {end_date:%d %b %Y}"
    )

    if data.empty:
        st.info(
            "No attendance records exist "
            "for the selected period."
        )
        return

    (
        strength,
        present,
        absent,
        percentage
    ) = calculate_metrics(data)

    metric_columns = st.columns(4)

    metric_columns[0].metric(
        "Attendance opportunities",
        strength
    )

    metric_columns[1].metric(
        "Present",
        present
    )

    metric_columns[2].metric(
        "Absent",
        absent
    )

    metric_columns[3].metric(
        "Attendance",
        f"{percentage:.2%}"
    )

    if period == "Day":
        preview = summarize_attendance(
            data,
            [
                "School_Code",
                "School"
            ]
        )

    else:
        preview = summarize_attendance(
            data,
            [
                "Date",
                "School_Code",
                "School"
            ]
        )

    st.dataframe(
        preview.style.format(
            {
                "Attendance_Percentage":
                    "{:.2%}"
            }
        ),
        use_container_width=True,
        hide_index=True
    )

    report_title = (
        f"{period} Attendance Report — "
        f"{selected_name}"
    )

    excel_data = make_excel_report(
        data,
        report_title,
        start_date,
        end_date
    )

    safe_period = (
        f"{start_date.isoformat()}_to_"
        f"{end_date.isoformat()}"
    )

    st.download_button(
        "Download Excel report",
        data=excel_data,
        file_name=(
            f"attendance_report_"
            f"{safe_period}.xlsx"
        ),
        mime=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        type="primary"
    )


# =========================================================
# CLASSES AND SECTIONS
# =========================================================

def render_class_setup(school_id):
    classes = get_school_classes(
        school_id,
        active_only=False
    )

    st.subheader(
        "Class and section configuration"
    )

    st.caption(
        "Disable the girls section if the school "
        "does not have girls. Previous records "
        "will remain unchanged."
    )

    if classes:

        class_options = {
            (
                f"{record['class_name']} "
                f"({'Active' if record['active'] else 'Inactive'})"
            ): record
            for record in classes
        }

        selected_label = st.selectbox(
            "Choose a class to edit",
            list(class_options.keys())
        )

        selected_class = class_options[
            selected_label
        ]

        with st.form(
            f"edit_class_{selected_class['id']}"
        ):

            class_name = st.text_input(
                "Class name",
                value=selected_class[
                    "class_name"
                ]
            )

            sort_order = st.number_input(
                "Display order",
                min_value=0,
                value=int(
                    selected_class[
                        "sort_order"
                    ]
                ),
                step=1
            )

            first_column, second_column = (
                st.columns(2)
            )

            boys_enabled = (
                first_column.checkbox(
                    "Boys section",
                    value=bool(
                        selected_class[
                            "boys_enabled"
                        ]
                    )
                )
            )

            boys_strength = (
                first_column.number_input(
                    "Boys strength",
                    min_value=0,
                    value=int(
                        selected_class[
                            "boys_strength"
                        ]
                    ),
                    step=1
                )
            )

            girls_enabled = (
                second_column.checkbox(
                    "Girls section",
                    value=bool(
                        selected_class[
                            "girls_enabled"
                        ]
                    )
                )
            )

            girls_strength = (
                second_column.number_input(
                    "Girls strength",
                    min_value=0,
                    value=int(
                        selected_class[
                            "girls_strength"
                        ]
                    ),
                    step=1
                )
            )

            active = st.checkbox(
                "Class is active",
                value=bool(
                    selected_class["active"]
                )
            )

            save_changes = (
                st.form_submit_button(
                    "Save class changes",
                    type="primary"
                )
            )

        if save_changes:

            if not class_name.strip():
                st.error(
                    "Class name is required."
                )

            elif (
                not boys_enabled
                and not girls_enabled
            ):
                st.error(
                    "Enable at least one section."
                )

            else:
                try:
                    with db_connection() as con:

                        con.execute(
                            """
                            UPDATE school_classes
                            SET
                                class_name = ?,
                                sort_order = ?,
                                boys_enabled = ?,
                                girls_enabled = ?,
                                boys_strength = ?,
                                girls_strength = ?,
                                active = ?
                            WHERE id = ?
                              AND school_id = ?
                            """,
                            (
                                class_name.strip(),
                                int(sort_order),
                                int(boys_enabled),
                                int(girls_enabled),

                                (
                                    int(boys_strength)
                                    if boys_enabled
                                    else 0
                                ),

                                (
                                    int(girls_strength)
                                    if girls_enabled
                                    else 0
                                ),

                                int(active),
                                selected_class["id"],
                                school_id
                            )
                        )

                    st.success(
                        "Class configuration updated."
                    )

                    st.rerun()

                except sqlite3.IntegrityError:
                    st.error(
                        "A class with this name "
                        "already exists."
                    )

    st.subheader("Add another class")

    with st.form(
        f"add_class_{school_id}"
    ):

        new_class_name = st.text_input(
            "New class name"
        )

        new_sort_order = st.number_input(
            "Display order",
            min_value=0,
            value=len(classes) + 1,
            step=1
        )

        first_column, second_column = (
            st.columns(2)
        )

        new_boys_enabled = (
            first_column.checkbox(
                "Boys section",
                value=True,
                key=f"new_boys_{school_id}"
            )
        )

        new_boys_strength = (
            first_column.number_input(
                "Boys strength",
                min_value=0,
                value=0,
                step=1,
                key=(
                    f"new_boys_strength_"
                    f"{school_id}"
                )
            )
        )

        new_girls_enabled = (
            second_column.checkbox(
                "Girls section",
                value=True,
                key=f"new_girls_{school_id}"
            )
        )

        new_girls_strength = (
            second_column.number_input(
                "Girls strength",
                min_value=0,
                value=0,
                step=1,
                key=(
                    f"new_girls_strength_"
                    f"{school_id}"
                )
            )
        )

        add_class = st.form_submit_button(
            "Add class"
        )

    if add_class:

        if not new_class_name.strip():
            st.error(
                "Enter a class name."
            )

        elif (
            not new_boys_enabled
            and not new_girls_enabled
        ):
            st.error(
                "Enable at least one section."
            )

        else:
            try:
                with db_connection() as con:

                    con.execute(
                        """
                        INSERT INTO school_classes(
                            school_id,
                            class_name,
                            sort_order,
                            boys_enabled,
                            girls_enabled,
                            boys_strength,
                            girls_strength,
                            active
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, 1)
                        """,
                        (
                            school_id,
                            new_class_name.strip(),
                            int(new_sort_order),
                            int(new_boys_enabled),
                            int(new_girls_enabled),

                            (
                                int(new_boys_strength)
                                if new_boys_enabled
                                else 0
                            ),

                            (
                                int(new_girls_strength)
                                if new_girls_enabled
                                else 0
                            )
                        )
                    )

                st.success("Class added.")
                st.rerun()

            except sqlite3.IntegrityError:
                st.error(
                    "A class with this name "
                    "already exists."
                )


# =========================================================
# AEO SCHOOL AND LOGIN MANAGEMENT
# =========================================================

def render_aeo_setup():
    st.title(
        "Schools, Usernames and Classes"
    )

    schools = get_active_schools()

    school_options = {
        (
            f"{school['code']} — "
            f"{school['name']}"
        ): school
        for school in schools
    }

    selected_label = st.selectbox(
        "Select school",
        list(school_options.keys())
    )

    selected_school = school_options[
        selected_label
    ]

    school_user = get_school_user(
        selected_school["id"]
    )

    st.subheader(
        "School name and username"
    )

    with st.form(
        f"school_details_"
        f"{selected_school['id']}"
    ):

        new_school_name = st.text_input(
            "School name",
            value=selected_school["name"]
        )

        new_username = st.text_input(
            "School login username",
            value=(
                school_user["username"]
                if school_user
                else ""
            )
        )

        save_school_details = (
            st.form_submit_button(
                "Update school and username",
                type="primary"
            )
        )

    if save_school_details:

        clean_school_name = (
            new_school_name.strip()
        )

        clean_username = (
            new_username.strip()
        )

        if not clean_school_name:
            st.error(
                "School name cannot be empty."
            )

        elif len(clean_username) < 3:
            st.error(
                "The username must contain "
                "at least 3 characters."
            )

        elif " " in clean_username:
            st.error(
                "The username cannot contain spaces."
            )

        else:
            try:
                with db_connection() as con:

                    con.execute(
                        """
                        UPDATE schools
                        SET name = ?
                        WHERE id = ?
                        """,
                        (
                            clean_school_name,
                            selected_school["id"]
                        )
                    )

                    con.execute(
                        """
                        UPDATE users
                        SET
                            username = ?,
                            display_name = ?
                        WHERE school_id = ?
                          AND role = 'SCHOOL'
                        """,
                        (
                            clean_username,
                            f"{clean_school_name} User",
                            selected_school["id"]
                        )
                    )

                st.success(
                    "School name and username updated."
                )

                st.rerun()

            except sqlite3.IntegrityError:
                st.error(
                    "This username is already "
                    "being used by another account."
                )

    with st.expander(
        "Reset this school's password"
    ):

        with st.form(
            f"reset_password_"
            f"{selected_school['id']}"
        ):

            new_password = st.text_input(
                "New temporary password",
                type="password"
            )

            confirm_password = st.text_input(
                "Confirm temporary password",
                type="password"
            )

            reset_password = (
                st.form_submit_button(
                    "Reset password"
                )
            )

        if reset_password:

            if new_password != confirm_password:
                st.error(
                    "Passwords do not match."
                )

            elif len(new_password) < 8:
                st.error(
                    "The password must contain "
                    "at least 8 characters."
                )

            else:
                with db_connection() as con:

                    con.execute(
                        """
                        UPDATE users
                        SET
                            password_hash = ?,
                            must_change_password = 1
                        WHERE school_id = ?
                          AND role = 'SCHOOL'
                        """,
                        (
                            hash_password(
                                new_password
                            ),
                            selected_school["id"]
                        )
                    )

                st.success(
                    "Temporary password reset. "
                    "The school must change it "
                    "after login."
                )

    render_class_setup(
        selected_school["id"]
    )


# =========================================================
# USER ACCOUNT
# =========================================================

def render_account(user):
    st.title("My Account")

    st.write(
        f"**Username:** {user['username']}"
    )

    st.write(
        f"**Role:** {user['role']}"
    )

    if user.get("school_name"):
        st.write(
            f"**School:** "
            f"{user['school_name']}"
        )

    st.subheader("Change username")

    with st.form(
        "account_username_change"
    ):

        new_username = st.text_input(
            "New username",
            value=user["username"]
        )

        username_password = st.text_input(
            "Current password to confirm",
            type="password"
        )

        change_username_button = (
            st.form_submit_button(
                "Change username"
            )
        )

    if change_username_button:

        successful, message = (
            change_own_username(
                user["id"],
                username_password,
                new_username
            )
        )

        if successful:
            st.session_state.user[
                "username"
            ] = new_username.strip()

            st.success(message)
            st.rerun()

        else:
            st.error(message)

    st.subheader("Change password")

    with st.form(
        "account_password_change"
    ):

        current_password = st.text_input(
            "Current password",
            type="password"
        )

        new_password = st.text_input(
            "New password",
            type="password"
        )

        confirm_password = st.text_input(
            "Confirm new password",
            type="password"
        )

        change_password_button = (
            st.form_submit_button(
                "Change password"
            )
        )

    if change_password_button:

        if new_password != confirm_password:
            st.error(
                "The new passwords do not match."
            )

        else:
            successful, message = change_password(
                user["id"],
                current_password,
                new_password
            )

            if successful:
                st.success(message)

            else:
                st.error(message)


# =========================================================
# MAIN NAVIGATION
# =========================================================

def render_application(user):
    st.sidebar.title("Roznamcha")

    st.sidebar.write(
        f"Signed in as "
        f"**{user['display_name']}**"
    )

    if user["role"] == "SCHOOL":

        st.sidebar.caption(
            user["school_name"]
        )

        pages = [
            "Dashboard",
            "Daily Attendance",
            "Reports",
            "Classes",
            "Account"
        ]

    else:
        st.sidebar.caption(
            "AEO — all schools"
        )

        pages = [
            "Dashboard",
            "Reports",
            "Schools & Classes",
            "Account"
        ]

    selected_page = st.sidebar.radio(
        "Menu",
        pages
    )

    if st.sidebar.button(
        "Sign out",
        use_container_width=True
    ):
        st.session_state.pop(
            "user",
            None
        )

        st.rerun()

    if selected_page == "Dashboard":

        if user["role"] == "AEO":
            render_aeo_dashboard()

        else:
            render_school_dashboard(user)

    elif selected_page == "Daily Attendance":
        render_attendance_entry(user)

    elif selected_page == "Reports":
        render_reports(user)

    elif selected_page == "Classes":
        st.title("Classes and Sections")

        render_class_setup(
            user["school_id"]
        )

    elif selected_page == "Schools & Classes":
        render_aeo_setup()

    elif selected_page == "Account":
        render_account(user)


# =========================================================
# RUN APPLICATION
# =========================================================

initialize_database()

if "user" not in st.session_state:
    render_login()

else:
    current_user = st.session_state.user

    if current_user["must_change_password"]:
        render_required_password_change(
            current_user
        )

    else:
        render_application(
            current_user
        )